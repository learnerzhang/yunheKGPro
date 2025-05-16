import pandas as pd
import datetime
from . import yautils
import os
#import yautils
import openpyxl
from datetime import datetime, timedelta
from openpyxl import load_workbook
from langchain.vectorstores import FAISS
import requests
import logging
import json

logger = logging.getLogger(__name__)

def getYuAnParamPath(ctype, mydate):
    """
        获取对应的参数路径
    """
    if ctype == 0:
        return f"data/yuan_data/0/plans/HHZXY_api_data_{mydate}.json"
    elif ctype == 1:
        return f"data/yuan_data/1/plans/XLDQX_api_data_{mydate}.json"
    elif ctype == 2:
        return f"data/yuan_data/2/plans/XLDTSTS_api_data_{mydate}.json"
    elif ctype == 3:
        return f"data/yuan_data/3/plans/HHXQ_api_data_{mydate}.json"
    elif ctype == 4:
        return f"data/yuan_data/4/plans/YLH_api_data_{mydate}.json"
    return "data/yuan_data/0/plans/HHZXY_api_data_2023-07-23.json"


def getYuAnName(ctype, mydate):
    """
    生产标准的预案名称
    """
    format_name = f"{mydate}黄河中下游防汛调度预案" 
    if ctype == 1:
        format_name = f"{mydate}小浪底秋汛防汛调度预案"
    elif ctype == 2:
        format_name = f"{mydate}小浪底调水调沙防汛调度预案"
    elif ctype == 3:
        format_name = f"{mydate}黄河汛情及水库调度方案单"
    elif ctype == 4:
        format_name = f"{mydate}伊洛河流域防汛调度初步方案"
    return format_name


def img2base64(imgpath):
    import base64
    with open(imgpath, 'rb') as f:
        data = f.read()
        encoded_string = base64.b64encode(data)
        return encoded_string.decode('utf-8')

def excel_to_html_with_merged_cells(file_path):
    """
    将 Excel 文件转换为 HTML 表格，保留合并单元格，并格式化数值和时间
    """
    # 加载 Excel 文件
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # 创建 HTML 表格
    html_table = '<table class="dataframe">\n'

    # 遍历每一行
    for row in sheet.iter_rows():
        html_table += '<tr>\n'
        for cell in row:
            # 检查单元格是否被合并
            is_merged = False
            for merged_range in sheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    is_merged = True
                    # 如果是合并区域的第一个单元格，设置 rowspan 和 colspan
                    if cell.coordinate == merged_range.start_cell.coordinate:
                        value = cell.value
                        if isinstance(value, (int, float)):
                            value = round(value, 2)  # 保留两位小数
                            # 添加单位并确保单位格式正确
                            cell_value = sheet.cell(row=1, column=cell.column).value
                            if cell_value is not None and "流量" in cell_value:
                                unit = "m³/s"
                            else:
                                unit = ""
                            value = f"{value} {unit}"
                        elif isinstance(value, datetime):
                            value = value.replace(microsecond=0)  # 去除微秒
                            if value.second >= 30:
                                value = (value + timedelta(minutes=1)).replace(second=0)  # 四舍五入到整点
                            else:
                                value = value.replace(second=0)  # 直接舍去秒数
                            value = value.strftime('%Y-%m-%d %H:%M')  # 格式化为年月日 时分
                        html_table += f'<th rowspan="{merged_range.size["rows"]}" colspan="{merged_range.size["columns"]}">{value}</th>\n'
                    break
            if not is_merged:
                # 普通单元格
                value = cell.value
                if isinstance(value, (int, float)):
                    value = round(value, 2)  # 保留两位小数
                    # 添加单位并确保单位格式正确
                    cell_value = sheet.cell(row=1, column=cell.column).value
                    if cell_value is not None and "流量" in cell_value:
                        unit = "m³/s"
                    else:
                        unit = ""
                    value = f"{value} {unit}"
                elif isinstance(value, datetime):
                    value = value.replace(microsecond=0)  # 去除微秒
                    if value.second >= 30:
                        value = (value + timedelta(minutes=1)).replace(second=0)  # 四舍五入到整点
                    else:
                        value = value.replace(second=0)  # 直接舍去秒数
                    value = value.strftime('%Y-%m-%d %H:%M')  # 格式化为年月日 时分
                if cell.row == 1 or cell.column == 1:
                    html_table += f'<th>{value}</th>\n'
                else:
                    html_table += f'<td>{value}</td>\n'
        html_table += '</tr>\n'
    html_table += '</table>'
    return html_table

def paraHtml(text):
    paraHtmlText = "<p>" + text +  "</p>"
    return paraHtmlText


def divHtml(text):
    paraHtmlText = "<div style='text-align: center;'>" + text +  "</div>"
    return paraHtmlText

def bold_left_align(text):
    return f"<div style='text-align: left;'><strong>{text}</strong></div>"


def pd2HtmlCSS():
    return """"""
    css = """
        <style>
            .table-container {
                width: 100%;
                margin: 20px;
                overflow-x: auto;
            }
            .dataframe {
                font-family: Arial, sans-serif;
                border-collapse: collapse;
                width: 100%;
            }
            .dataframe thead {
                background-color: #6c7ae0 !important
                color: #000000;
            }
            .dataframe th, .dataframe td {
                padding: 12px 15px;
                text-align: left;
                border: 1px solid #ddd;
            }
            .dataframe tbody tr {
                border-bottom: 1px solid #ddd;
            }
            .dataframe tbody tr:nth-child(even) {
                background-color: #f3f3f3;
            }
            .dataframe tbody tr:hover {
                background-color: #f1f1f1;
            }
            .dataframe tbody tr:last-child {
                border-bottom: 2px solid #009879;
            }
        </style>
        """
    return css





def smx_sk(shuiku_shuiwei: dict, shuiwenzhan_liuliang: dict):
    sw = shuiku_shuiwei.get("三门峡", {}).get('level', 0)
    hyk_liuliang = shuiwenzhan_liuliang.get("花园口",{}).get('flow', 0)
    print("三门峡SK", hyk_liuliang, sw)
    """
    三门峡
    :return:
    """
    if hyk_liuliang <= 4500:
        return "视潼关站来水来沙情况，原则上按敞泄运用"
    elif hyk_liuliang <= 8000:
        return "视潼关站来水来沙情况，原则上按敞泄运用"
    elif hyk_liuliang <= 10000:
        return "视潼关站来水来沙情况，原则上按敞泄运用"
    elif hyk_liuliang <= 22000:
        return "原则上敞泄运用，视水库蓄水及来水情况适时控泄"
    else:
        return "适时控制运用"


# def xld_sk(shuiku_shuiwei: dict, shuiwenzhan_liuliang: dict):
#     sw = shuiku_shuiwei.get("小浪底", {}).get('level', 0)
#     hyk_liuliang = shuiwenzhan_liuliang.get("花园口", {}).get('flow', 0)
#     tongguan_liuliang = shuiwenzhan_liuliang.get("潼关", {}).get('flow', 0)
#     rate_p = float(tongguan_liuliang) / (float(hyk_liuliang) + 0.001)
#     print("小浪底SK", hyk_liuliang,tongguan_liuliang, sw)
#     """
#     小浪底库
#     :return:
#     """
#     if hyk_liuliang <= 4500:
#         return "适时调节水沙，按控制花园口站流量不大于4500m³/s的原则泄洪。西霞院水库配合小浪底水库泄洪排沙"
#     elif hyk_liuliang <= 8000:
#         return "原则上按控制花园口站4500m³/s方式运用。若洪水主要来源于三门峡以上，视来水来沙及水库淤积情况，适时按进出库平衡方式运用。控制水库最高运用水位不超过254m。西霞院水库配合小浪底水库泄洪排沙"
#     elif hyk_liuliang <= 10000:
#         if rate_p > 0.6:
#             return "小浪底：原则上按进出库平衡方式运用。西霞院水库配合小浪底水库泄洪排沙"
#         else:
#             return "小浪底：视下游汛情，适时按控制花园口站不大于8000m³/s的方式运用。西霞院水库配合小浪底水库泄洪排沙"
#     elif hyk_liuliang <= 22000:
#         if rate_p > 0.6:
#             return "小浪底：按控制花园口站10000m³/s方式运用。西霞院水库配合小浪底水库泄洪排沙"
#         else:
#             if hyk_liuliang - tongguan_liuliang < 9000:
#                 return "小浪底：按控制花园口站10000m³/s方式运用。西霞院水库配合小浪底水库泄洪排沙"
#             else:
#                 return "小浪底：按不大于1000m³/s（发电流量）下泄。西霞院水库配合小浪底水库泄洪排沙"
#     else:
#         # 潼关 花园口 流量大于60%
#         if rate_p > 0.6:
#             if sw < 273.5:
#                 return "小浪底：按控制花园口站10000m³/s方式运用。西霞院水库配合小浪底水库泄洪排沙"
#             else:
#                 return "小浪底：按进出库平衡或敞泄运用。西霞院水库配合小浪底水库泄洪排沙"
#         else:
#             return "小浪底：按控制花园口站10000m³/s方式运用。西霞院水库配合小浪底水库泄洪排沙"


def lh_sk(shuiku_shuiwei: dict, shuiwenzhan_liuliang: dict):
    """
    陆浑水库
    :return:
    """
    sw = shuiku_shuiwei.get("陆浑", {}).get('level', 0)
    hyk_liuliang = shuiwenzhan_liuliang.get("花园口", {}).get('flow', 0)
    print("小浪底SK", hyk_liuliang, sw)

    if hyk_liuliang <= 4500:
        if sw < 321.5:
            return "按控制下泄流量不大于1000m³/s方式运用"
        else:
            return "按进出库平衡或敞泄运用"
    elif hyk_liuliang <= 8000:
        if sw < 321.5:
            return "按控制下泄流量不大于1000m³/s方式运用"
        else:
            return "按进出库平衡或敞泄运用"
    elif hyk_liuliang <= 10000:
        if sw < 321.5:
            return "按控制下泄流量不大于1000m³/s方式运用"
        else:
            return "按进出库平衡或敞泄运用"
    elif hyk_liuliang <= 22000:
        if hyk_liuliang <= 12000:
            if sw < 321.5:
                return "按控制下泄流量不大于1000m³/s方式运用"
            else:
                return "按进出库平衡或敞泄运用"
        else:
            if sw < 323:
                return "按不超过发电流量控泄"
            else:
                return "按进出库平衡或敞泄运用"
    else:
        if sw < 323:
            return "按不超过发电流量控泄"
        else:
            return "按进出库平衡或敞泄运用"


def gx_sk(shuiku_shuiwei: dict, shuiwenzhan_liuliang: dict):
    """
    故县水库
    :return:
    """
    sw = shuiku_shuiwei.get("故县", {}).get('level', 0)
    hyk_liuliang = shuiwenzhan_liuliang.get("花园口", {}).get('flow', 0)
    print("故县SK", hyk_liuliang, sw)

    if hyk_liuliang <= 4500:
        if sw < 542.04:
            return "按控制下泄流量不大于1000m³/s方式运用"
        else:
            return "按进出库平衡或敞泄运用"
    elif hyk_liuliang <= 8000:
        if sw < 542.04:
            return "按控制下泄流量不大于1000m³/s方式运用"
        else:
            return "按进出库平衡或敞泄运用"
    elif hyk_liuliang <= 10000:
        if sw < 542.04:
            return "按控制下泄流量不大于1000m³/s方式运用"
        else:
            return "按进出库平衡或敞泄运用"
    elif hyk_liuliang <= 22000:
        if hyk_liuliang <= 12000:
            if sw < 542.04:
                return "按控制下泄流量不大于1000m³/s方式运用"
            else:
                return "按进出库平衡或敞泄运用"
        else:
            if sw < 546.84:
                return "按不超过发电流量控泄"
            else:
                return "按进出库平衡或敞泄运用"
    else:
        if sw < 546.84:
            return "按不超过发电流量控泄"
        else:
            return "按进出库平衡或敞泄运用"


def hkc_sk(shuiku_shuiwei: dict, shuiwenzhan_liuliang: dict):
    """
    河口村水库
    :return:
    """
    sw = shuiku_shuiwei.get("河口村", {}).get('level', 0)
    hyk_liuliang = shuiwenzhan_liuliang.get("花园口", {}).get('flow', 0)
    print("河口村SK", hyk_liuliang, sw)
    if hyk_liuliang <= 4500:
        if sw < 254.5:
            return "按控制武陟不大于2000m³/s方式运用。张峰水库适时配合河口村水库拦洪运用"
        elif sw < 285.43:
            return "按控制武陟不大于4000m³/s方式运用。张峰水库适时配合河口村水库拦洪运用"
        else:
            return "按进出库平衡方式运用。张峰水库适时配合河口村水库拦洪运用"
    elif hyk_liuliang <= 8000:
        if sw < 254.5:
            return "按控制武陟不大于2000m³/s方式运用。张峰水库适时配合河口村水库拦洪运用"
        elif sw < 285.43:
            return "按控制武陟不大于4000m³/s方式运用。张峰水库适时配合河口村水库拦洪运用"
        else:
            return "按进出库平衡方式运用。张峰水库适时配合河口村水库拦洪运用"
    elif hyk_liuliang <= 10000:
        if sw < 254.5:
            return "按控制武陟不大于2000m³/s方式运用。张峰水库适时配合河口村水库拦洪运用"
        elif sw < 285.43:
            return "按控制武陟不大于4000m³/s方式运用。张峰水库适时配合河口村水库拦洪运用"
        else:
            return "按进出库平衡方式运用。张峰水库适时配合河口村水库拦洪运用"
    elif hyk_liuliang <= 22000:
        if hyk_liuliang <= 12000:
            if sw < 254.5:
                return "按控制武陟不大于2000m³/s方式运用。张峰水库适时配合河口村水库拦洪运用"
            elif sw < 285.43:
                return "按控制武陟不大于4000m³/s方式运用。张峰水库适时配合河口村水库拦洪运用"
            else:
                return "按进出库平衡方式运用。张峰水库适时配合河口村水库拦洪运用"
        else:
            if sw < 254.5:
                return "关闭所有泄流设施。张峰水库适时配合河口村水库拦洪运用"
            elif sw < 285.43:
                return "尽可能按控制武陟不大于4000m³/s方式运用。张峰水库适时配合河口村水库拦洪运用"
            else:
                return "按进出库平衡方式运用。张峰水库适时配合河口村水库拦洪运用"
    else:
        if sw < 254.5:
            return "关闭所有泄流设施。张峰水库适时配合河口村水库拦洪运用"
        elif sw < 285.43:
            return "尽可能按控制武陟不大于4000m³/s方式运用。张峰水库适时配合河口村水库拦洪运用"
        else:
            return "按进出库平衡方式运用。张峰水库适时配合河口村水库拦洪运用"
        


def get_xunqi():
    day = datetime.date.today()
    m = day.month
    if m in [7, 8]:  # 前汛期
        return 1
    if m in [9, 10]:  # 后汛期
        return 2
    return 0  # 其他其


def xld_sk(sw=300, a=0.5):
    """小浪底水库"""
    xq = get_xunqi()
    if xq == 1:
        if sw >= 281:
            return "超坝顶高程{}m".format(round(sw - 281, 2))
        elif 281 - a <= sw < 281:
            return "低于坝顶高程{}m".format(round(281 - sw, 2))
        elif 275 <= sw < 281 - a:
            return "超校核洪水位（防洪高水位）{}m".format(round(sw - 275, 2))
        elif 275 - a <= sw < 275:
            return "低于校核洪水位（防洪高水位）{}m".format(round(275 - sw, 2))
        elif 274 <= sw < 275 - a:
            return "超设计洪水位{}m".format(round(sw - 274, 2))
        elif 273.5 <= sw < 274:
            return "超历史最高水位{}m".format(round(sw - 273.5, 2))
        elif 273 - a <= sw < 273.5:
            return "低于历史最高水位{}m".format(round(273.5 - sw, 2))
        elif 235 <= sw < 273 - a:
            return "超汛限水位{}m".format(round(sw - 235, 2))
        elif 235 - a <= sw < 235:
            return "低于汛限水位{}m".format(round(235 - sw, 2))
        elif 230 <= sw < 235 - a:
            return "超死水位{}m".format(round(sw - 230, 2))
        else:
            return "低于死水位{}m".format(round(230 - sw, 2))
    elif xq == 2:
        if sw >= 281:
            return "超坝顶高程{}m".format(round(sw - 281, 2))
        elif 281 - a <= sw < 281:
            return "低于坝顶高程{}m".format(round(281 - sw, 2))
        elif 275 <= sw < 281 - a:
            return "超校核洪水位（防洪高水位）{}m".format(round(sw - 275, 2))
        elif 275 - a <= sw < 275:
            return "低于校核洪水位（防洪高水位）{}m".format(round(275 - sw, 2))
        elif 274 <= sw < 275 - a:
            return "超设计洪水位{}m".format(round(sw - 274, 2))
        elif 273.5 <= sw < 274:
            return "超历史最高水位{}m".format(round(sw - 273.5, 2))
        elif 273 - a <= sw < 273.5:
            return "低于历史最高水位{}m".format(round(273.5 - sw, 2))
        elif 248 <= sw < 273 - a:
            return "超汛限水位{}m".format(round(sw - 235, 2))
        elif 248 - a <= sw < 248:
            return "低于汛限水位{}m".format(round(235 - sw, 2))
        elif 230 <= sw < 248 - a:
            return "超死水位{}m".format(round(sw - 230, 2))
        else:
            return "低于死水位{}m".format(round(230 - sw, 2))
    return ""


def dph_sk(sw=60, a=0.5):
    """
    东平湖
    :param sw:
    :param a:
    :return:
    """
    xq = get_xunqi()
    if xq == 1:
        if sw >= 43.22:
            return "超防洪运用水位{}m".format(sw - 43.22)
        elif 43.22 - a <= sw < 43.22:
            return "低于防洪运用水位{}m".format(43.22 - sw)
        elif 40.72 <= sw < 43.22 - a:
            return "超汛限水位{}m".format(sw - 40.72)
        else:
            return "低于汛限水位{}m".format(40.72 - sw)
    elif xq == 2:
        if sw >= 43.22:
            return "超防洪运用水位{}m".format(sw - 43.22)
        elif 43.22 - a <= sw < 43.22:
            return "低于防洪运用水位{}m".format(43.22 - sw)
        elif 41.72 <= sw < 43.22 - a:
            return "超汛限水位{}m".format(sw - 41.72)
        else:
            return "低于汛限水位{}m".format(41.72 - sw)


def smx_sjt_sk(sw=410.0, a=0.5):
    """
    三门峡（史家摊）
    :param sw:
    :param a:
    :return:
    """
    if sw >= 351.65:
        return "超坝顶高程{}m".format(round(sw - 351.65, 2))
    elif 351.65 - a <= sw < 351.65:
        return "低于坝顶高程{}m".format(round(351.65 - sw, 2))
    elif 335 <= sw < 351.65 - a:
        return "超校核洪水位{}m".format(round(sw - 335, 2))
    elif 335 - a <= sw < 335:
        return "超校核洪水位{}m".format(round(335 - sw, 2))
    elif 333.65 <= sw < 335 - a:
        return "超防洪运用水位{}m".format(round(sw - 333.65, 2))
    elif 333.65 - a <= sw < 333.65:
        return "低于防洪运用水位{}m".format(round(333.65 - sw, 2))
    elif 331.23 <= sw < 333.65 - a:
        return "超历史最高水位{}m".format(round(sw - 331.23, 2))
    elif 331.23 - a <= sw < 331.23:
        return "低于历史最高水位{}m".format(round(331.23 - sw, 2))
    elif 318 <= sw < 332.58 - a:
        return "超人员紧急转移水位{}m".format(round(sw - 318, 2))
    elif 318 - a <= sw < 318:
        return "低于人员紧急转移水位{}m".format(round(318 - sw, 2))
    elif 305 <= sw < 318 - a:
        return "超汛限水位{}m".format(round(sw - 305, 2))
    else:
        return "低于汛限水位{}m".format(round(305 - sw, 2))


def hkc_sk(sw=300, a=0.5):
    """
    河口村水库
    :param sw:
    :param a:
    :return:
    """
    xq = get_xunqi()
    if xq == 1:
        if sw >= 288.5:
            return "超坝顶高程{}m".format(round(sw - 288.5, 2))
        elif 288.5 - a <= sw < 288.5:
            return "低于坝顶高程{}m".format(round(288.5 - sw, 2))
        elif 285.43 <= sw < 288.5 - a:
            return "超校核洪水位（防洪高水位）{}m".format(round(sw - 285.43, 2))
        elif 285.43 - a <= sw < 285.43:
            return "低于校核洪水位（防洪高水位）{}m".format(round(285.43 - sw, 2))
        elif 283 <= sw < 285.43 - a:
            return "超人员紧急转移水位{}m".format(round(sw - 283, 2))
        elif 283 - a <= sw < 283:
            return "低于人员紧急转移水位{}m".format(round(283 - sw, 2))
        elif 279.89 <= sw < 283 - a:
            return "超历史最高水位{}m".format(round(sw - 279.89, 2))
        elif 279.89 - a <= sw < 279.89:
            return "低于历史最高水位{}m".format(round(279.89 - sw, 2))
        elif 238 <= sw < 279.89 - a:
            return "超汛限水位{}m".format(round(sw - 238, 2))
        elif 238 - a <= sw < 238:
            return "低于汛限水位{}m".format(round(238 - sw, 2))
        elif 225 <= sw < 238 - a:
            return "超死水位{}m".format(round(sw - 225, 2))
        else:
            return "低于死水位{}m".format(round(225 - sw, 2))
    elif xq == 2:
        if sw >= 288.5:
            return "超坝顶高程{}m".format(round(sw - 288.5, 2))
        elif 288.5 - a <= sw < 288.5:
            return "低于坝顶高程{}m".format(round(288.5 - sw, 2))
        elif 285.43 <= sw < 288.5 - a:
            return "超校核洪水位（防洪高水位）{}m".format(round(sw - 285.43, 2))
        elif 285.43 - a <= sw < 285.43:
            return "低于校核洪水位（防洪高水位）{}m".format(round(285.43 - sw, 2))
        elif 283 <= sw < 285.43 - a:
            return "超人员紧急转移水位{}m".format(round(sw - 283, 2))
        elif 283 - a <= sw < 283:
            return "低于人员紧急转移水位{}m".format(round(283 - sw, 2))
        elif 279.89 <= sw < 283 - a:
            return "超历史最高水位{}m".format(round(sw - 279.89, 2))
        elif 279.89 - a <= sw < 279.89:
            return "低于历史最高水位{}m".format(round(279.89 - sw, 2))
        elif 275 <= sw < 279.89 - a:
            return "超汛限水位{}m".format(round(sw - 275, 2))
        elif 275 - a <= sw < 275:
            return "低于汛限水位{}m".format(round(275 - sw, 2))
        elif 225 <= sw < 275 - a:
            return "超死水位{}m".format(round(sw - 225, 2))
        else:
            return "低于死水位{}m".format(round(225 - sw, 2))
    return ""


def gx_sk(sw=600, a=0.5):
    """
    故县水库
    :param sw:
    :param a:
    :return:
    """
    xq = get_xunqi()
    if xq == 1:
        if sw >= 551.84:
            return "超坝顶高程{}m".format(round(sw - 551.84, 2))
        elif 551.84 - a <= sw < 551.84:
            return "低于坝顶高程{}m".format(round(551.84 - sw, 2))
        elif 549.86 <= sw < 551.48 - a:
            return "超校核洪水位（防洪高水位）{}m".format(round(sw - 549.86, 2))
        elif 549.86 - a <= sw < 549.86:
            return "低于校核洪水位（防洪高水位）{}m".format(round(549.86 - sw, 2))
        elif 547.39 <= sw < 549.86 - a:
            return "超设计洪水位{}m".format(round(sw - 547.39, 2))
        elif 547.39 - a <= sw < 547.39:
            return "低于设计洪水位{}m".format(round(547.39 - sw, 2))
        elif 546.84 <= sw < 547.39 - a:
            return "超防洪高水位（蓄洪限制水位）{}m".format(round(sw - 546.84, 2))
        elif 546.84 - a <= sw < 546.84:
            return "低于防洪高水位（蓄洪限制水位）{}m".format(round(546.84 - sw, 2))
        elif 543.04 <= sw < 546.84 - a:
            return "超移民水位{}m".format(round(sw - 543.04, 2))
        elif 543.04 - a <= sw < 543.04:
            return "低于移民水位{}m".format(round(543.04 - sw, 2))
        elif 537.75 <= sw < 543.04 - a:
            return "超历史最高水位{}m".format(round(sw - 537.75, 2))
        elif 537.75 - a <= sw < 537.75:
            return "低于历史最高水位{}m".format(round(537.75 - sw, 2))
        elif 533.64 <= sw < 537.75 - a:
            return "超征地水位{}m".format(round(sw - 533.64, 2))
        elif 533.64 - a <= sw < 533.64:
            return "低于征地水位{}m".format(round(533.64 - sw, 2))
        elif 527.3 <= sw < 533.64 - a:
            return "超汛限水位{}m".format(round(sw - 527.3, 2))
        elif 527.3 - a <= sw < 527.3:
            return "超汛限水位{}m".format(round(527.3 - sw, 2))
        elif 498.84 <= sw < 527.3 - a:
            return "超死水位{}m".format(round(sw - 498.84, 2))
        else:
            return "低于死水位{}m".format(round(498.84 - sw, 2))
    elif xq == 2:
        if sw >= 551.84:
            return "超坝顶高程{}m".format(round(sw - 551.84, 2))
        elif 551.84 - a <= sw < 551.84:
            return "低于坝顶高程{}m".format(round(551.84 - sw, 2))
        elif 549.86 <= sw < 551.48 - a:
            return "超校核洪水位（防洪高水位）{}m".format(round(sw - 549.86, 2))
        elif 549.86 - a <= sw < 549.86:
            return "低于校核洪水位（防洪高水位）{}m".format(round(549.86 - sw, 2))
        elif 547.39 <= sw < 549.86 - a:
            return "超设计洪水位{}m".format(round(sw - 547.39, 2))
        elif 547.39 - a <= sw < 547.39:
            return "低于设计洪水位{}m".format(round(547.39 - sw, 2))
        elif 546.84 <= sw < 547.39 - a:
            return "超防洪高水位（蓄洪限制水位）{}m".format(round(sw - 546.84, 2))
        elif 546.84 - a <= sw < 546.84:
            return "低于防洪高水位（蓄洪限制水位）{}m".format(round(546.84 - sw, 2))
        elif 543.04 <= sw < 546.84 - a:
            return "超移民水位{}m".format(round(sw - 543.04, 2))
        elif 543.04 - a <= sw < 543.04:
            return "低于移民水位{}m".format(round(543.04 - sw, 2))
        elif 537.75 <= sw < 543.04 - a:
            return "超历史最高水位{}m".format(round(sw - 537.75, 2))
        elif 537.75 - a <= sw < 537.75:
            return "低于历史最高水位{}m".format(round(537.75 - sw, 2))
        elif 534.3 <= sw < 537.75 - a:
            return "超汛限水位{}m".format(round(sw - 534.3, 2))
        elif 534.3 - a <= sw < 534.3:
            return "低于汛限水位{}m".format(round(534.3 - sw, 2))
        elif 533.64 <= sw < 534.3 - a:
            return "超征地水位{}m".format(round(sw - 533.64, 2))
        elif 533.64 - a <= sw < 533.64:
            return "低于征地水位{}m".format(round(533.64 - sw, 2))
        elif 498.84 <= sw < 533.64 - a:
            return "超死水位{}m".format(round(sw - 498.84, 2))
        else:
            return "低于死水位{}m".format(round(498.84 - sw, 2))
    return ""


def lhbs_sk(sw=380, a=0.5):
    """
    陆浑坝上
    :param sw:
    :param a:
    :return:
    """
    sw = round(sw, 2)
    xq = get_xunqi()
    if xq == 1:
        if sw >= 333:
            return "超坝顶高程{}m".format(round(sw - 333, 2))
        elif 333 - a <= sw < 333:
            return "低于坝顶高程{}m".format(round(333 - sw, 2))
        elif 331.8 <= sw < 333 - a:
            return "超校核洪水位（防洪高水位）{}m".format(round(sw - 331.8, 2))
        elif 331.8 - a <= sw < 331.8:
            return "低于校核洪水位（防洪高水位）{}m".format(round(331.8 - sw, 2))
        elif 327.5 <= sw < 331.8 - a:
            return "超设计洪水位{}m".format(round(sw - 327.5, 2))
        elif 327.5 - a <= sw < 327.5:
            return "低于设计洪水位{}m".format(round(327.5 - sw, 2))
        elif 325 <= sw < 327.5 - a:
            return "超移民水位{}m".format(round(sw - 325, 2))
        elif 325 - a <= sw < 325:
            return "低于移民水位{}m".format(round(325 - sw, 2))
        elif 323 <= sw < 325 - a:
            return "超蓄洪限制水位{}m".format(round(sw - 323, 2))
        elif 323 - a <= sw < 323:
            return "低于蓄洪限制水位{}m".format(round(323 - sw, 2))
        elif 320.91 <= sw < 323 - a:
            return "超历史最高水位{}m".format(round(sw - 320.91, 2))
        elif 320.91 - a <= sw < 320.91:
            return "低于历史最高水位{}m".format(round(320.91 - sw, 2))
        elif 319.5 <= sw < 320.91 - a:
            return "超征地水位{}m".format(round(sw - 319.5, 2))
        elif 319.5 - a <= sw < 319.5:
            return "低于征地水位{}m".format(round(319.5 - sw, 2))
        elif 317 <= sw < 319.5 - a:
            return "超汛限水位{}m".format(round(sw - 317, 2))
        elif 317 - a <= sw < 317:
            return "低于汛限水位{}m".format(round(317 - sw, 2))
        elif 298 <= sw < 317 - a:
            return "超死水位{}m".format(round(sw - 298, 2))
        else:
            return "低于死水位{}m".format(round(498.84 - sw, 2))
        
    elif xq == 2:
        if sw >= 333:
            return "超坝顶高程{}m".format(round(sw - 333, 2))
        elif 333 - a <= sw < 333:
            return "低于坝顶高程{}m".format(round(333 - sw, 2))
        elif 331.8 <= sw < 333 - a:
            return "超校核洪水位（防洪高水位）{}m".format(round(sw - 331.8, 2))
        elif 331.8 - a <= sw < 331.8:
            return "低于校核洪水位（防洪高水位）{}m".format(round(331.8 - sw, 2))
        elif 327.5 <= sw < 331.8 - a:
            return "超设计洪水位{}m".format(round(sw - 327.5, 2))
        elif 327.5 - a <= sw < 327.5:
            return "低于设计洪水位{}m".format(round(327.5 - sw, 2))
        elif 325 <= sw < 327.5 - a:
            return "超移民水位{}m".format(round(sw - 325, 2))
        elif 325 - a <= sw < 325:
            return "低于移民水位{}m".format(round(325 - sw, 2))
        elif 323 <= sw < 325 - a:
            return "超蓄洪限制水位{}m".format(round(sw - 323, 2))
        elif 323 - a <= sw < 323:
            return "低于蓄洪限制水位{}m".format(round(323 - sw, 2))
        elif 320.91 <= sw < 323 - a:
            return "超历史最高水位{}m".format(round(sw - 320.91, 2))
        elif 320.91 - a <= sw < 320.91:
            return "低于历史最高水位{}m".format(round(320.91 - sw, 2))
        elif 319.5 <= sw < 320.91 - a:
            return "超征地水位{}m".format(round(sw - 319.5, 2))
        elif 319.5 - a <= sw < 319.5:
            return "低于征地水位{}m".format(round(319.5 - sw, 2))
        elif 317.5 <= sw < 319.5 - a:
            return "超汛限水位{}m".format(round(sw - 317.5, 2))
        elif 317.5 - a <= sw < 317.5:
            return "低于汛限水位{}m".format(round(317.5 - sw, 2))
        elif 298 <= sw < 317.5 - a:
            return "超死水位{}m".format(round(sw - 298, 2))
        else:
            return "低于死水位{}m".format(round(498.84 - sw, 2))
    return ""


import re
def text_table(text):
    """
    从输入文本中提取黄河下游工程数据，并转换为结构化数据。

    参数:
        text (str): 输入文本。

    返回:
        list: 包含各河段及总计数据的字典列表。
    """
    # 定义正则表达式提取数据
    pattern = re.compile(
        r"(\w+段)累计有(\d+)处工程(\d+)道坝出险(\d+)次，抢险用石([\d.]+)万方，耗资([\d.]+)万元"
    )
    # 提取数据
    matches = pattern.findall(text)

    # 转换为目标格式
    data = []
    total = {
        "河段": "总计",
        "工程数量": 0,
        "出险坝数": 0,
        "出险次数": 0,
        "抢险用石（万方）": 0.0,
        "耗资（万元）": 0.0
    }

    for match in matches:
        segment_data = {
            "河段": match[0],
            "工程数量": int(match[1]),
            "出险坝数": int(match[2]),
            "出险次数": int(match[3]),
            "抢险用石（万方）": float(match[4]),
            "耗资（万元）": float(match[5])
        }
        data.append(segment_data)

        # 计算总计
        total["工程数量"] += segment_data["工程数量"]
        total["出险坝数"] += segment_data["出险坝数"]
        total["出险次数"] += segment_data["出险次数"]
        total["抢险用石（万方）"] += segment_data["抢险用石（万方）"]
        total["耗资（万元）"] += segment_data["耗资（万元）"]

    # 添加总计行
    data.append(total)

    return data

def extract_shuiku_data(text):
    """
    从输入文本中提取水库名称和运用方式，并转换为结构化数据。

    参数:
        text (str): 输入文本。

    返回:
        list: 包含水库名称和运用方式的字典列表。
    """
    # 定义正则表达式提取数据
    # pattern = re.compile(
    #     r"([\u4e00-\u9fa5]+水库)：(.+?)\n"
    # )
    pattern = re.compile(
        r"([\u4e00-\u9fa5]+水库)：(.+?)(\n|$)"
    )
    # 提取数据
    matches = pattern.findall(text)

    # 转换为目标格式
    data = []
    for match in matches:
        reservoir_data = {
            "水库": match[0],
            "调度原则": match[1].strip()
        }
        data.append(reservoir_data)

    return data


def extract_shuiku_data_jianyi(text):
    """
    从输入文本中提取水库名称和运用方式，并转换为结构化数据。

    参数:
        text (str): 输入文本。

    返回:
        list: 包含水库名称和运用方式的字典列表。
    """
    # 定义正则表达式提取数据
    # pattern = re.compile(
    #     r"([\u4e00-\u9fa5]+水库)：(.+?)\n"
    # )
    pattern = re.compile(
        r"([\u4e00-\u9fa5]+水库)：(.+?)(\n|$)"
    )
    # 提取数据
    matches = pattern.findall(text)

    # 转换为目标格式
    data = []
    for match in matches:
        reservoir_data = {
            "水库": match[0],
            "调度建议": match[1].strip()
        }
        data.append(reservoir_data)

    return data

def process_outflow(data, timestamps):
    # 记录每次第一个不同的数的索引
    unique_indices = []
    unique_values = []

    # 找到每个不同出库流量的索引
    for index in range(len(data)):
        if index == 0 or data[index] != data[index - 1]:
            unique_indices.append(index)
            unique_values.append(data[index])

    # 组织输出信息
    output = []
    for i in range(len(unique_indices)):
        if i < len(unique_indices) - 1:  # 如果不是最后一个索引
            duration = unique_indices[i + 1] - unique_indices[i]  # 当前值到下一个值的差
        else:  # 如果是最后一个索引
            duration = len(data) - unique_indices[i]  # 最后一个值到结束
        # if i == 0:
        #     duration = unique_indices[i + 1] - unique_indices[i]  # 第一个值到第二个值的差
        # elif i < len(unique_indices) - 1:
        #     duration = unique_indices[i + 1] - unique_indices[i]  # 当前值到下一个值的差
        # else:
        #     duration = len(data) - unique_indices[i]  # 最后一个值到结束

        # 计算具体的小时数
        hours = duration * 2
        start_time = timestamps[unique_indices[i]]
        if i < len(unique_indices) - 1:
            end_time = timestamps[unique_indices[i + 1]]
        else:
            end_time = timestamps[-1]  # 最后一个值的结束时间为 timestamps 的最后一个元素
        output.append(f"{start_time}起按{unique_values[i]}m³/s泄放 {hours}个小时,至{end_time}结束")
    res = "，".join(output) + "。"
    return res

def generate_ddjy_v1(file_path):
    """
    解析 Excel 表格数据。

    参数:
        file_path (str): Excel 文件路径。

    返回:
        list: 包含解析后的数据的字典列表。
    """
    # 读取 Excel 文件
    df = pd.read_excel(file_path, header=None)
    # 处理表头
    # 获取前两行作为表头信息
    first_row = df.iloc[0, :].fillna('')  # 第一行
    second_row = df.iloc[1, :].fillna('')  # 第二行
    # 合并表头信息
    header = []
    for i in range(len(first_row)):
        prefix = str(first_row[i]).strip()
        suffix = str(second_row[i]).strip()
        if prefix and suffix:
            header.append(f"{prefix} {suffix}")
        elif prefix:
            header.append(prefix)
        else:
            header.append(suffix)
    # 设置列名
    df.columns = header
    df = df.drop([0, 1, 2])  # 删除前3行
    # 动态匹配列名
    time_column = None
    for col in header:
        if "月.日" in col:
            time_column = col
            break
    if time_column is None:
        raise ValueError("未找到时间列")
    tongguan_column = None
    for col in header:
        if "潼关" in col and "预报流量" in col:
            tongguan_column = col
            break
    if tongguan_column is None:
        raise ValueError("未找到潼关预报流量列")
    sanmenxia_inflow = None
    for col in header:
        if "三门峡" in col and "入库流量" in col:
            sanmenxia_inflow = col
            break
    if sanmenxia_inflow is None:
        raise ValueError("未找到三门峡入库流量列")
    #print(df.iloc[:,1])
    # 解析数据
    time_stamps,smx_ckll, xld_ckll, lh_ckll,gx_ckll, hkc_ckll=[],[],[],[],[],[]
    for _, row in df.iterrows():
        time_value = row[time_column]
        formatted_time = ""
        if isinstance(time_value, pd.Timestamp):
            formatted_time = time_value.strftime('%Y-%m-%d %H')  # 格式化为 "YYYY-MM-DD HH:MM:SS"
        elif isinstance(time_value, str):
            try:
                time_obj = pd.to_datetime(time_value)
                formatted_time = time_obj.strftime('%Y-%m-%d %H')  # 格式化为 "YYYY-MM-DD HH:MM:SS"
            except ValueError:
                formatted_time = time_value  # 如果解析失败，保持原样
        else:
            formatted_time = str(time_value)  # 如果不是时间格式，转为字符串
            if '.' in formatted_time:  # 检查是否包含毫秒
                formatted_time = formatted_time.split('.')[0]  # 去除毫秒部分
            else:
                formatted_time = formatted_time  # 如果没有毫秒，保持
        time_stamps.append(formatted_time)
        smx_ckll.append(row.iloc[3])
        xld_ckll.append(row.iloc[7])
        lh_ckll.append(row.iloc[11])
        gx_ckll.append(row.iloc[15])
        hkc_ckll.append(row.iloc[19])
    # print(time_stamps, "长度：",len(time_stamps))
    # print(smx_ckll, "长度：",len(smx_ckll))
    # print(xld_ckll, "长度：",len(xld_ckll))
    # print(lh_ckll, "长度：",len(lh_ckll))
    # print(gx_ckll, "长度：",len(gx_ckll))
    # print(hkc_ckll, "长度：",len(hkc_ckll))
    smx_ddjy = process_outflow(smx_ckll,time_stamps)
    xld_ddjy = process_outflow(xld_ckll,time_stamps)
    lh_ddjy = process_outflow(lh_ckll,time_stamps)
    gx_ddjy = process_outflow(gx_ckll,time_stamps)
    hkc_ddjy = process_outflow(hkc_ckll,time_stamps)
    data = "三门峡水库："+smx_ddjy+"\n小浪底水库："+xld_ddjy+"\n陆浑水库："+lh_ddjy+"\n故县水库："+gx_ddjy+"\n河口村水库："+hkc_ddjy
    return data

def generate_ddjy(file_path):
    """
    解析 Excel 表格数据。

    参数:
        file_path (str): Excel 文件路径。

    返回:
        list: 包含解析后的数据的字典列表。
    """
    # 读取 Excel 文件
    skMapData, swMapData, date_list = yautils.excel_to_dict(file_path)
    smx_ckll,xld_ckll,lh_ckll,gx_ckll,hkc_ckll = yautils.skddjy(file_path)
    smx_ddjy = process_outflow(smx_ckll,date_list)
    xld_ddjy = process_outflow(xld_ckll,date_list)
    lh_ddjy = process_outflow(lh_ckll,date_list)
    gx_ddjy = process_outflow(gx_ckll,date_list)
    hkc_ddjy = process_outflow(hkc_ckll,date_list)
    data = "三门峡水库："+smx_ddjy+"\n小浪底水库："+xld_ddjy+"\n陆浑水库："+lh_ddjy+"\n故县水库："+gx_ddjy+"\n河口村水库："+hkc_ddjy
    return data

def yujingdengji(shuiku_shuiwei: dict, shuiwenzhan_liuliang: dict):
    print("预警等级")
    """
    TODO 条件
    :param smx_sw:
    :param lh_sw:
    :param xld_sw:
    :param gx_sw:
    :param dph_sw:
    :param hkc_sw:
    :param swz_ll:
    :return:
    """
    smx_sw = shuiku_shuiwei.get("三门峡", {}).get('level', 0)
    xld_sw = shuiku_shuiwei.get("小浪底", {}).get('level', 0)
    lh_sw = shuiku_shuiwei.get("陆浑", {}).get('level', 0)
    gx_sw = shuiku_shuiwei.get("故县", {}).get('level', 0)
    hkc_sw = shuiku_shuiwei.get("河口村", {}).get('level', 0)
    dph_sw = shuiku_shuiwei.get("东平湖", {}).get('level', 0)

    knh_ll = shuiwenzhan_liuliang.get("唐乃亥", {}).get('flow', 0)
    lz_ll = shuiwenzhan_liuliang.get("兰州", {}).get('flow', 0)
    szs_ll = shuiwenzhan_liuliang.get("石嘴山", {}).get('flow', 0)
    lm_ll = shuiwenzhan_liuliang.get("龙门", {}).get('flow', 0)
    tg_ll = shuiwenzhan_liuliang.get("潼关", {}).get('flow', 0)
    hx_ll = shuiwenzhan_liuliang.get("华县", {}).get('flow', 0)
    hyk_ll = shuiwenzhan_liuliang.get("花园口", {}).get('flow', 0)
    gc_ll = shuiwenzhan_liuliang.get("高村", {}).get('flow', 0)
    wl_ll = shuiwenzhan_liuliang.get("武陟", {}).get('flow', 0)
    act_flag, lev = None, None#huangwei_yujing_rec()
    print(act_flag, lev)
    if (act_flag and lev == 'Ⅰ') or lh_sw >= 331.8 or hkc_sw >= 285.43 or dph_sw >= 43.22 or gx_sw >= 549.86 or xld_sw >= 275 or smx_sw >= 335 or knh_ll >= 5000 or lz_ll >= 6500 or szs_ll >= 5500 or lm_ll >= 18000 or tg_ll >= 15000 or hyk_ll >= 15000 or hx_ll >= 8000 or wl_ll >= 4000:
        print("# 启动防汛一级应急响应")
        return """按照《黄河防汛抗旱应急预案》，启动一级应急响应，响应行动如下：（1）黄河防总总指挥或常务副总指挥坐镇指挥黄河抗洪工作，主持抗洪抢险会商会，研究部署抗洪抢险工作。视情与相关省区进行异地会商。（2）根据会商意见，黄河防总办公室向相关省区防指通报关于启动防汛一级应急响应的命令及黄河汛情，对防汛工作提出要求，并向黄河防总总指挥报告。黄河防总向国家防总、水利部报告有关情况，为国家防总和水利部提供调度参谋意见，请求加强对黄河抗洪抢险指导，动员社会力量支援黄河抗洪抢险救灾。（3）黄河防总办公室各成员单位按照黄委防御大洪水职责分工和机构设置上岗到位，全面开展工作，各职能组充实人员。黄委全体职工全力投入抗洪抢险工作。水情测报组滚动进行洪水预测预报，每日至少制作发布气象水情预报 3 次，每日至少提供12 次干支流重要测站监测信息，情况紧急时根据需要加密测报；综合调度组根据预报滚动计算水利工程调度方案，做好干流及重要支流水库调度和东平湖、北金堤滞洪区运用的分析研判；宣传组适时举行新闻发布会，向社会报道黄河抗洪抢险动态，做好新闻宣传工作。（4）黄河防总根据汛情需要，及时增派司局级领导带队的工作组、专家组赶赴现场，指导抗洪抢险救灾工作。（5）根据各地抗洪抢险需要，黄河防总按程序调度黄委防汛物资、黄河机动抢险队支援抗洪抢险，必要时请求国家防总调动流域内外抢险队、物资支援黄河抗洪抢险。（6）有关省区防汛抗旱指挥机构的主要负责同志主持会商，动员部署防汛工作；按照权限组织调度水工程；根据预案转移安置危险地区群众，组织强化巡堤查险和堤防防守，及时控制险情；增派工作组、专家组赴一线指导防汛工作；受灾地区的各级防汛抗旱指挥机构负责人、成员单位负责人，应按照职责到分管的区域组织指挥防汛工作，或驻点具体帮助重灾区做好防汛工作；可按照预案和程序适时请调人民解放军和武警部队支援黄河抗洪抢险；将工作情况上报省区人民政府及黄河防总。根据汛情，相关县级以上人民政府防汛抗旱指挥部宣布进入紧急防汛期，动员一切社会力量投入黄河抗洪抢险"""
    if (act_flag and lev == 'Ⅱ') or lh_sw >= 327.5 or hkc_sw >= 285.43 or dph_sw >= 43.22 - 0.5 or gx_sw >= 547.39 or xld_sw >= 274 or smx_sw >= 335 or knh_ll >= 4000 or lz_ll >= 5000 or szs_ll >= 4000 or lm_ll >= 12000 or tg_ll >= 10000 or hyk_ll >= 8000 or hx_ll >= 6000 or wl_ll >= 3000:
        print("# 启动防汛二级应急响应")
        return """按照《黄河防汛抗旱应急预案》，启动二级应急响应，响应行动如下：（1）黄河防总总指挥或常务副总指挥坐镇指挥黄河抗洪工作，主持抗洪抢险会商会，研究部署抗洪抢险工作。视情与相关省区进行异地会商。（2）根据会商意见，黄河防总办公室向相关省区防指通报关于启动防汛二级应急响应的命令及黄河汛情，对防汛工作提出要求，并向黄河防总总指挥报告。黄河防总向国家防总、水利部报告有关情况，为国家防总和水利部提供调度参谋意见，请求加强对黄河抗洪抢险指导。（3）黄河防总办公室各成员单位按照黄委防御大洪水职责分工和机构设置上岗到位，全面开展工作。黄委全体职工做好随时投入抗洪抢险工作的准备。（4）黄河防总实时掌握雨情、水情、汛情（凌情）、工情、险情、灾情动态。水情测报组滚动进行洪水预测预报，每日至少制作发布气象水情预报 2 次，每日至少提供 6 次干支流重要测站监测信息，情况紧急时根据需要加密测报；综合调度组根据预报滚动计算水利工程调度方案，做好干流及重要支流水库调度和东平湖滞洪区运用的分析研判；宣传组定期举行新闻发布会，向社会公布黄河抗洪抢险动态。（5）黄河防总办公室根据汛情需要，及时派出司局级领导带队的工作组、专家组赶赴现场，检查、指导抗洪抢险救灾工作，核实汛情灾情。（6）根据各地抗洪抢险需要，黄河防总办公室按程序调度黄委防汛物资、黄河机动抢险队支援抗洪抢险。（7）有关省区防汛抗旱指挥机构负责同志主持会商，具体安排防汛工作；按照权限组织调度水工程；根据预案做好巡堤查险、抗洪抢险、群众转移安置等抗洪救灾工作，派出工作组、专家组赴一线指导防汛工作；将防汛工作情况上报省级人民政府主要负责同志、国家防总及黄河防总。按照预案和程序适时请调人民解放军和武警部队支援黄河抗洪抢险。根据汛情，相关县级以上人民政府防汛抗旱指挥部宣布进入紧急防汛期。"""
    if (act_flag and lev == 'Ⅲ') or lh_sw >= 319.5 or hkc_sw >= 285.43 or dph_sw >= 43.22 or gx_sw >= 549.86 or smx_sw >= 335 or knh_ll >= 3000 or lz_ll >= 4000 or szs_ll >= 3000 or lm_ll >= 8000 or tg_ll >= 8000 or hyk_ll >= 6000 or hx_ll >= 4000 or wl_ll >= 2000:
        # 启动防汛三级应急响应
        print("# 启动防汛三级应急响应")
        return """按照《黄河防汛抗旱应急预案》，启动三级应急响应，响应行动如下：（1）黄河防总秘书长主持防汛会商会，研究部署抗洪抢险工作。视情与相关省区进行异地会商。（2）根据会商意见，黄河防总办公室向相关省区防指通报关于启动防汛三级应急响应的命令及黄河汛情，对防汛工作提出要求，并向黄河防总总指挥、常务副总指挥报告。黄河防总向国家防总、水利部报告有关情况，为国家防总和水利部提供调度参谋意见，请求加强对黄河抗洪抢险指导。（3）黄河防总办公室各成员单位按照黄委防御大洪水职责分工和机构设置上岗到位，全面开展工作。水情测报组滚动进行洪水预测预报，每日至少制作发布气象水情预报 1 次，每日至少提供 3 次（8 时、14 时、20 时）干支流重要测站监测信息，情况紧急时根据需要加密测报；综合调度组根据预报滚动计算水利工程调度方案，做好干流及重要支流水库调度；宣传组加强黄河抗洪抢险宣传。（4）黄河防总办公室根据汛情需要，及时派出工作组、专家组赶赴现场，检查、指导抗洪抢险救灾工作，核实汛情灾情。黄委防汛物资、黄河机动抢险队支援抗洪抢险。（5）根据各地抗洪抢险需要，黄河防总办公室按程序调度黄委防汛物资、黄河机动抢险队支援抗洪抢险。（6）有关省区防汛抗旱指挥机构负责同志主持会商，具体安排防汛工作；按照权限组织调度水工程；根据预案做好巡堤查险、抗洪抢险、群众转移安置等抗洪救灾工作，派出工作组、专家组赴一线指导防汛工作；将防汛工作情况上报省级人民政府分管负责同志和黄河防总。可按照预案和程序适时请调人民解放军和武警部队支援黄河抗洪抢险。在省级主要媒体及新媒体平台发布防汛抗旱有关情况。"""
    if (act_flag and lev == 'IV') or lh_sw >= 331.8 or hkc_sw >= 285.43 or dph_sw >= 43.22 or gx_sw >= 549.86 or smx_sw >= 335 or knh_ll >= 2500 or lz_ll >= 2500 or szs_ll >= 2000 or lm_ll >= 5000 or tg_ll >= 5000 or hyk_ll >= 4000 or hx_ll >= 2500 or wl_ll >= 1000:
        # 启动防汛四级应急响应
        print("# 启动防汛四级应急响应")
        return """按照《黄河防汛抗旱应急预案》，启动四级应急响应，响应行动如下：（1）黄河防总秘书长主持会商，研究部署抗洪抢险工作，确定运行机制。响应期间，根据汛情发展变化，受黄河防总秘书长委托，可由黄河防总办公室副主任主持会商，并将情况报黄河防总秘书长。（2）根据会商意见，黄河防总办公室向相关省区防指通报关于启动防汛四级应急响应的命令及黄河汛情，对防汛工作提出要求，并向国家防办、水利部报告有关情况，必要时向黄河防总总指挥、常务副总指挥报告。（3）黄河防总办公室成员单位人员坚守工作岗位，加强防汛值班值守。按照黄委防御大洪水职责分工和机构设置，综合调度、水情测报和工情险情组等人员上岗到位。其余成员单位按照各自职责做好技术支撑、通信保障、后勤及交通保障，加强宣传报道。水情测报组及时分析天气形势并结合雨水情发展态势，做好雨情、水情、沙情的预测预报，加强与水利部信息中心、黄河流域气象中心、省区气象水文部门会商研判，每日至少制作发布气象水情预报 1 次，每日至少提供 2 次（8 时、20 时）干支流重要测站监测信息，情况紧急时根据需要加密测报。（4）黄委按照批准的洪水调度方案，结合当前汛情做好水库等水工程调度，监督指导地方水行政主管部门按照调度权限做好水工程调度。（5）黄河防总办公室根据汛情需要，及时派出工作组、专家组赶赴现场，检查、指导抗洪抢险救灾工作，核实汛情灾情。（6）有关省区防汛抗旱指挥机构负责同志主持会商，具体安排防汛工作；按照权限组织调度水工程；按照预案做好辖区内巡堤查险、抗洪抢险、群众转移安置等抗洪救灾工作，必要时请调解放军和武警部队、民兵参加重要堤段、重点工程的防守或突击抢险；派出工作组、专家组赴一线指导防汛工作；将防汛工作情况上报省级人民政府和黄河防总办公室。"""
    print("""预警等级: 按照《黄河防汛抗旱应急预案》，当前无预警""")
    return """按照《黄河防汛抗旱应急预案》，当前无预警"""



def search_fragpacks(query, top_k=5):
    from yunheKGPro.settings import embedding
    """
    根据 query 检索知识片段，并返回分数最高的文本片段。

    参数:
        query (str): 查询字符串。
        top_k (int): 返回的相似片段数量，默认为 1。

    返回:
        分数最高的文本片段（str）。
        如果向量库不存在或未检索到结果，返回 None。
    """
    # 构建知识库目录路径
    kgdir = f"../data/knowledges/264c159aa9e1dd91e31ab9f38f3d4f9c"
    if not os.path.exists(kgdir):
        os.makedirs(kgdir)
    # 构建向量库文件路径
    index_path = os.path.join(kgdir, "faiss.index")
    # 如果向量库不存在，返回 None
    if not os.path.exists(index_path):
        print("❌ 向量库不存在")
        return None
    # 加载 embedding 模型和向量库
    print("✅ 加载 embedding 模型")
    db = FAISS.load_local(index_path, embedding, allow_dangerous_deserialization=True)
    print("✅ 已加载现有向量库")
    # 执行向量检索
    fragpacks = db.similarity_search_with_score(query, top_k)

    # 如果没有检索到结果，返回 None
    if not fragpacks:
        print("❌ 未检索到任何结果")
        return None
    # 找到分数最高的结果
    highest_score_doc = max(fragpacks, key=lambda x: x[1])  # 按分数排序，取最高分
    return highest_score_doc[0].page_content

BASE_API_URL= "http://wt.hxyai.cn/fx/"#"http://10.4.158.35:8070/"#
def get_rainfall_data_hour(basin=None, start_time=None, end_time=None):
    """
    获取实时雨量数据（带完整错误处理和JWT鉴权）

    参数:
        basin: 流域编码 (如"RH01")
        start_time: 开始时间 (如"2024-03-01 08:00:00")
        end_time: 结束时间 (如"2024-03-02 08:00:00")

    返回:
        (status_code, response_data) 元组
    """
    # 构造请求参数
    params = {}
    if basin:
        params["stcd"] = basin
    if start_time:
        params["startDate"] = start_time
    if end_time:
        params["endDate"] = end_time

    try:
        # 使用提供的JWT token
        auth_token = oauth_login()

        # 添加超时和请求头
        response = requests.get(
            url=f"{BASE_API_URL}/rainfall/hourrth/getRainfall",
            params=params,
            headers={
                "Accept": "application/json",
                "Authorization": f"{auth_token}"
            },
            timeout=10  # 10秒超时
        )
        response.raise_for_status()  # 自动检查4xx/5xx错误
        return response.status_code, response.json()

    except requests.exceptions.Timeout:
        return 408, {"error": "请求超时"}
    except requests.exceptions.RequestException as e:
        return 500, {"error": f"请求失败: {str(e)}"}
    except ValueError:  # JSON解析错误
        return 500, {"error": "返回数据格式异常"}


def get_rainfall_data_day(auth_token,basin=None, start_time=None, end_time=None):
    """
    获取实时雨量数据（带完整错误处理和JWT鉴权）

    参数:
        basin: 流域编码 (如"RH01")
        start_time: 开始时间 (格式: "2024-03-01")，若不传则默认为昨天
        end_time: 结束时间 (格式: "2024-03-02")，若不传则默认为今天

    返回:
        (status_code, response_data) 元组
    """
    # 构造请求参数
    params = {}
    # 如果当前时间小于8点，则开始时间是昨天8点，否则是今天8点
    now = datetime.now()
    if now.hour < 8:
        start_default = (now.replace(hour=8, minute=0, second=0, microsecond=0))- timedelta(days=1)
    else:
        start_default = now.replace(hour=8, minute=0, second=0, microsecond=0)

    end_default = start_default + timedelta(days=1)#now#
    # 设置默认时间范围（今天和昨天）
    # today = datetime.now().strftime("%Y-%m-%d")
    # yesterday = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
    # # 处理时间参数
    # params["endDate"] = end_default #end_time if end_time else today
    # params["startDate"] = start_default #start_time if start_time else yesterday
    params["endTime"] = end_time if end_time else end_default.strftime("%Y-%m-%d %H:%M:%S")
    params["startTime"] = start_time if start_time else start_default.strftime("%Y-%m-%d %H:%M:%S")
    print("params:",params)
    if basin:
        params["stcd"] = basin
    try:
        # 使用提供的JWT token
        auth_token = auth_token#oauth_login()
        url = f"{BASE_API_URL}/rainfall/hourrth/getRainfall"
        # 添加超时和请求头
        response = requests.get(
            url=f"{BASE_API_URL}/rainfall/hourrth/getRainfall",#dayrt
            params=params,
            headers={
                "Accept": "application/json",
                "Authorization": f"{auth_token}"
            },
            timeout=10  # 10秒超时
        )
        response.raise_for_status()  # 自动检查4xx/5xx错误
        print("url:",url)
        return response.status_code, response.json()

    except requests.exceptions.Timeout:
        return 408, {"error": "请求超时"}
    except requests.exceptions.RequestException as e:
        return 500, {"error": f"请求失败: {str(e)}"}
    except ValueError:  # JSON解析错误
        return 500, {"error": "返回数据格式异常"}

def get_max_rainfall_station(data):
    """
    从降雨量数据中找出降雨量最大的站点

    参数:
        data: 列表，包含各个站点的数据字典，格式见示例

    返回:
        dict: 包含最大降雨量站点的 'stnm'(站名), 'lgtd'(经度), 'lttd'(纬度)
        或 None（如果所有站点降雨量都为0）
    """
    if data == []:
        return None
    max_rainfall = -1
    max_station = None
    yiluo_stations = [station for station in data
                      if str(station['stcd']).startswith('416')]
    for station in yiluo_stations:
        rainfall = station.get("rf", 0)
        if rainfall > max_rainfall:
            max_rainfall = rainfall
            max_station = station

    # 如果所有站点的降雨量都是0，则返回None
    if max_rainfall == 0:
        return None

    return {
        "stnm": max_station["stnm"],
        "lgtd": max_station["lgtd"],
        "lttd": max_station["lttd"],
        "rf": max_station["rf"]
    }


def generate_rainfall_report_v1(response_data):
    """
    生成降雨量报告文本

    参数:
        response_data: API返回的JSON数据

    返回:
        降雨量报告文本
    """
    if not response_data or 'data' not in response_data:
        return "无有效降雨数据"

    # 筛选伊洛河流域站点(416开头)
    yiluo_stations = [station for station in response_data['data']
                      if str(station['stcd']).startswith('416')]

    if not yiluo_stations:
        return "伊洛河流域无降雨监测数据"

    # 统计降雨等级
    rainfall_stats = {
        '小雨': 0,
        '中雨': 0,
        '大雨': 0,
        '暴雨': 0,
        '大暴雨': 0,
        '特大暴雨': 0
    }

    max_rainfall = 0
    max_station = ""

    for station in yiluo_stations:
        rf = station['rf']

        if rf == 0:
            continue

        if rf < 10:
            rainfall_stats['小雨'] += 1
        elif 10 <= rf < 25:
            rainfall_stats['中雨'] += 1
        elif 25 <= rf < 50:
            rainfall_stats['大雨'] += 1
        elif 50 <= rf < 100:
            rainfall_stats['暴雨'] += 1
        elif 100 <= rf < 250:
            rainfall_stats['大暴雨'] += 1
        else:
            rainfall_stats['特大暴雨'] += 1

        # 记录最大降雨量
        if rf > max_rainfall:
            max_rainfall = rf
            max_station = station['stnm']

    # 如果没有降雨
    if all(count == 0 for count in rainfall_stats.values()):
        return f"{datetime.now().month}月{datetime.now().day}日，伊洛河流域无降雨"

    # 生成报告文本
    today = datetime.now()
    report_parts = [f"{today.month}月{today.day}日，伊洛河流域"]

    # 添加降雨描述
    rainfall_desc = []
    if rainfall_stats['小雨'] > 0:
        rainfall_desc.append("小雨")
    if rainfall_stats['中雨'] > 0:
        rainfall_desc.append("中雨")
    if rainfall_stats['大雨'] > 0:
        rainfall_desc.append("大雨")
    if rainfall_stats['暴雨'] > 0:
        rainfall_desc.append("暴雨")
    if rainfall_stats['大暴雨'] > 0:
        rainfall_desc.append("大暴雨")
    if rainfall_stats['特大暴雨'] > 0:
        rainfall_desc.append("特大暴雨")

    if rainfall_desc:
        report_parts.append("大部地区降" + "、".join(rainfall_desc))

    # 添加最大降雨点
    if max_rainfall > 0:
        report_parts.append(f"，最大点雨量{max_station}站{max_rainfall}毫米")

    return "".join(report_parts)

def get_ylh_rainfall():
    """
    获取伊洛河流域面雨量数据
    返回格式：字符串 "伊洛河流域面雨量XX毫米" 或错误信息字符串
    """
    # 计算时间范围（昨天8点到今天8点）
    now = datetime.now()
    end_time = now.replace(hour=8, minute=0, second=0, microsecond=0)
    start_time = end_time - timedelta(days=1)

    # 格式化时间参数
    params = {
        "startTime": start_time.strftime("%Y-%m-%d %H:%M:%S"),
        "endTime": end_time.strftime("%Y-%m-%d %H:%M:%S")
    }

    try:
        # 发送请求
        response = requests.get(
            #"http://10.4.158.35:8092/rfFace/YLH/RainAnalysis",#经过nginx转发的地址
            "http://10.4.158.36:9091/YLH/RainAnalysis",#原始地址
            params=params,
            #timeout=10
        )
        response.raise_for_status()

        # 解析响应数据
        data = response.json()
        if data.get("code") == "200":
            return f"伊洛河流域面雨量{data.get('meanrain', 0)}毫米"
        return data.get("message", "接口返回数据异常")

    except requests.exceptions.Timeout:
        return "请求超时，请稍后重试"
    except requests.exceptions.RequestException as e:
        return f"请求失败：{str(e)}"
    except (KeyError, ValueError) as e:
        return "数据解析错误"

def generate_rainfall_report(response_data):
    """
    生成降雨量报告文本(版本1)
    按照新的规则生成降雨报告:
    1. 优先判断大部降雨(≥50%站点)
    2. 其次判断局部降雨(≥20%站点)
    3. 再次判断个别地区降雨(>0%站点)
    4. 最后判断无降雨

    参数:
        response_data: API返回的JSON数据

    返回:
        降雨量报告文本
    """
    today = datetime.now()
    if not response_data or 'data' not in response_data:
        return f"{today.month}月{today.day}日，未获取到伊洛河流域降雨数据。"

    # 筛选伊洛河流域站点(416开头)
    yiluo_stations = [station for station in response_data['data']
                      if str(station['stcd']).startswith('416')]

    if not yiluo_stations:
        return f"{today.month}月{today.day}日，伊洛河流域无降雨。"#"伊洛河流域无降雨监测数据"

    # 统计降雨等级和站点总数
    rainfall_stats = {
        '小雨': 0,
        '中雨': 0,
        '大雨': 0,
        '暴雨': 0,
        '大暴雨': 0,
        '特大暴雨': 0
    }

    max_rainfall = 0
    max_station = ""
    total_stations = len(yiluo_stations)
    rain_stations = 0  # 有降雨的站点数

    for station in yiluo_stations:
        rf = station['rf']

        if rf == 0:
            continue

        rain_stations += 1

        if rf < 10:
            rainfall_stats['小雨'] += 1
        elif 10 <= rf < 25:
            rainfall_stats['中雨'] += 1
        elif 25 <= rf < 50:
            rainfall_stats['大雨'] += 1
        elif 50 <= rf < 100:
            rainfall_stats['暴雨'] += 1
        elif 100 <= rf < 250:
            rainfall_stats['大暴雨'] += 1
        else:
            rainfall_stats['特大暴雨'] += 1

        # 记录最大降雨量
        if rf > max_rainfall:
            max_rainfall = rf
            max_station = station['stnm']

    # 如果没有降雨
    if rain_stations == 0:
        today = datetime.now()
        return f"{today.month}月{today.day}日，伊洛河流域无降雨。"

    # 计算各雨型的站点比例
    ratios = {
        '小雨': rainfall_stats['小雨'] / total_stations,
        '中雨': rainfall_stats['中雨'] / total_stations,
        '大雨': rainfall_stats['大雨'] / total_stations,
        '暴雨': rainfall_stats['暴雨'] / total_stations,
        '大暴雨': rainfall_stats['大暴雨'] / total_stations,
        '特大暴雨': rainfall_stats['特大暴雨'] / total_stations
    }

    # 生成报告文本
    today = datetime.now()
    meanrain = get_ylh_rainfall()
    print("面雨量：",meanrain)
    report_parts = [f"{today.month}月{today.day}日"]
    if isinstance(meanrain, str) and "面雨量" in meanrain:
        report_parts.append(f"，{meanrain}")
    else:
        pass
    report_parts.append("，伊洛河流域")
    # 判断降雨范围
    # 1. 优先判断大部降雨(≥50%)
    # 大部暴雨
    if ratios['暴雨'] >= 0.5:
        report_parts.append("大部地区降暴雨")
    # 大部大雨
    elif ratios['大雨'] >= 0.5:
        report_parts.append("大部地区降大雨")
    # 大部中雨
    elif ratios['中雨'] >= 0.5:
        report_parts.append("大部地区降中雨")
    # 大部小雨
    elif ratios['小雨'] >= 0.5:
        report_parts.append("大部地区降小雨")
    # 大部大到暴雨
    elif (ratios['大雨'] + ratios['暴雨']) >= 0.5:
        report_parts.append("大部地区降大到暴雨")
    # 大部中到大雨
    elif (ratios['中雨'] + ratios['大雨']) >= 0.5:
        report_parts.append("大部地区降中到大雨")
    # 大部小到中雨
    elif (ratios['小雨'] + ratios['中雨']) >= 0.5:
        report_parts.append("大部地区降小到中雨")

    # 2. 其次判断局部降雨(≥20%)
    elif ratios['暴雨'] >= 0.2:
        report_parts.append("局部地区降暴雨")
    elif ratios['大雨'] >= 0.2:
        report_parts.append("局部地区降大雨")
    elif ratios['中雨'] >= 0.2:
        report_parts.append("局部地区降中雨")
    elif ratios['小雨'] >= 0.2:
        report_parts.append("局部地区降小雨")
    elif (ratios['大雨'] + ratios['暴雨']) >= 0.2:
        report_parts.append("局部地区降大到暴雨")
    elif (ratios['中雨'] + ratios['大雨']) >= 0.2:
        report_parts.append("局部地区降中到大雨")
    elif (ratios['小雨'] + ratios['中雨']) >= 0.2:
        report_parts.append("局部地区降小到中雨")

    # 3. 再次判断个别地区降雨(>0%)
    elif rainfall_stats['暴雨'] > 0:
        report_parts.append("个别地区降暴雨")
    elif rainfall_stats['大雨'] > 0:
        report_parts.append("个别地区降大雨")
    elif rainfall_stats['中雨'] > 0:
        report_parts.append("个别地区降中雨")
    elif rainfall_stats['小雨'] > 0:
        report_parts.append("个别地区降小雨")

    # 添加最大降雨点
    if max_rainfall > 0:
        report_parts.append(f"，最大点雨量{max_station}站{max_rainfall}毫米。")

    return "".join(report_parts)


def get_hydrometric_station(auth_token=None, station_code=None, ):
    """
    获取河道水文站基本信息

    参数:
        station_code: 水文站代码 (String, 可选)
        auth_token: JWT认证令牌 (String)

    返回:
        (status_code, response_data) 元组
    """
    auth_token = auth_token#oauth_login()

    try:
        headers = {
            "Accept": "application/json",
            "Authorization": f"{auth_token}"
        }

        params = {}
        if station_code:
            params["stcd"] = station_code
            params["stcd"] = station_code
        HYDROMETRIC_API_URL = f"{BASE_API_URL}/hydrometric/hourrt/listLatest"
        response = requests.get(
            HYDROMETRIC_API_URL,
            params=params,
            headers=headers,
            timeout=10
        )

        response.raise_for_status()
        return response.status_code, response.json()

    except requests.exceptions.Timeout:
        return 408, {"error": "请求超时"}
    except requests.exceptions.RequestException as e:
        return 500, {"error": f"请求失败: {str(e)}"}
    except ValueError:
        return 500, {"error": "返回数据格式异常"}

AUTH_TOKEN_PATH = "data/yuan_data/4/auth_token/auth_token.json"
def oauth_login_new(
    access_key: str =  "fxylh2",   #"fxylh",#
    secret_key: str = "899d657383d458a546bd80f1a0753263",#"656ed363fa5513bb9848b430712290b2",#
    user_type: int = 3
):
    url = f"{BASE_API_URL}/oauth/login"
    headers = {
        "Content-Type": "application/json",
        "Accept": "application/json"
    }
    payload = {
        "accessKey": access_key,
        "secretKey": secret_key,
        "userType": user_type
    }
    try:
        response = requests.post(url=url, headers=headers, data=json.dumps(payload), timeout=10)
        response.raise_for_status()
        data = response.json().get("data")
        if data:
            # 确保路径存在
            os.makedirs(os.path.dirname(AUTH_TOKEN_PATH), exist_ok=True)

            # 写入新的 token 到文件
            with open(AUTH_TOKEN_PATH, 'w') as f:
                json.dump({"auth_token": data}, f)

            return data

        else:
            # 如果没有 token 数据，尝试读取本地缓存
            return read_local_token()

    except (requests.exceptions.RequestException, ValueError) as e:
        # 请求失败或响应解析失败时尝试读取本地 token
        return read_local_token()


def read_local_token():
    """从本地文件读取缓存的 auth_token"""
    if os.path.exists(AUTH_TOKEN_PATH):
        try:
            with open(AUTH_TOKEN_PATH, 'r') as f:
                data = json.load(f)
                return data.get("auth_token")
        except Exception:
            return None
    return None

def oauth_login(
        access_key: str =  "fxylh2",#"fxylh",
        secret_key: str = "899d657383d458a546bd80f1a0753263",#"656ed363fa5513bb9848b430712290b2",
        user_type: int = 3
) :
    """
    """
    #url = "http://10.4.158.35:8070/oauth/login"
    url = f"{BASE_API_URL}oauth/login"
    headers = {"Content-Type": "application/json",
        "Accept": "application/json"}
    payload = {"accessKey": access_key, "secretKey": secret_key, "userType": user_type}
    try:
        response = requests.post(url=url,headers=headers,data=json.dumps(payload),timeout=10)
        response.raise_for_status()
        return response.json()['data']

    except requests.exceptions.RequestException as e:
        return {"code": 500,"data": None,"msg": f"请求失败: {str(e)}"}
    except ValueError:
        return {"code": 500,"data": None,"msg": "响应数据解析失败"}



def format_hydrometric_data_v1(auth_token=None,station_code=None):
    """
    获取并格式化水文站数据（仅返回YLH_HD列表中的站点）
    参数:
        station_code: 水文站代码 (可选)
    返回:
        (status_code, formatted_data) 元组
        formatted_data格式示例:
        {
            "hdsq": [
                {
                    "站名": "站名1",
                    "当日8时流量(m³/s)": 100.5,  # 仅8点数据有值，否则为None
                    "昨日日均流量": 95.2
                },
                ...
            ]
        }
    """
    # 首先获取原始数据
    auth_token = auth_token#oauth_login()
    status_code, raw_data = get_hydrometric_station(auth_token=auth_token,station_code=station_code)
    lyh_hd = ["卢氏", "长水", "宜阳", "白马寺", "黑石关", "东湾", "陆军", "龙门镇", "花园口"]
    # 如果请求不成功，直接返回原始结果
    if status_code != 200:
        return status_code, raw_data
    # 初始化格式化后的数据
    formatted_data = {"hdsq": []}
    try:
        # 处理数据
        if 'data' in raw_data and isinstance(raw_data['data'], list):
            for station in raw_data['data']:
                # 只处理416开头的站点且在YLH_HD列表中的站点
                if str(station.get('stcd', '')).startswith('416') and station.get('stnm') in lyh_hd:
                    # 获取昨日日均流量
                    hysta = station.get('stcd')
                    dayrt_status, dayrt_data = get_hydrometric_dayrt_list(auth_token=auth_token,hysta=hysta)
                    yesterday_avg_flow = None
                    if dayrt_status == 200 and dayrt_data.get('data') and len(dayrt_data['data']) > 0:
                        yesterday_avg_flow = dayrt_data['data'][0]["flow"]
                    # 检查时间是否为当日8点
                    timestamp = station.get('date')
                    eight_am_flow = None
                    if timestamp:
                        dt = datetime.fromtimestamp(timestamp / 1000)
                        now = datetime.now()
                        if dt.date() == now.date() and dt.hour == 8:  # 仅当日期为今日且时间为8点时记录流量
                            eight_am_flow = station.get('dstrvm')
                    # 构建格式化条目
                    formatted_entry = {
                        "站名": station.get('stnm', '未知站名'),
                        "当日8时流量(m³/s)": eight_am_flow,  # 非8点数据为None
                        "昨日日均流量": yesterday_avg_flow
                    }
                    formatted_data["hdsq"].append(formatted_entry)
        return status_code, formatted_data

    except Exception as e:
        return 500, {"error": f"数据处理错误: {str(e)}"}


def format_hydrometric_data(auth_token=None, station_code=None):
    """
    获取并格式化水文站数据（固定返回9个指定站点）
    修改：返回所有指定站点的数据，没有数据的站点设为None

    参数:
        auth_token: 认证令牌
        station_code: 水文站代码 (可选)
    返回:
        (status_code, formatted_data) 元组
        formatted_data格式示例:
        {
            "hdsq": [
                {
                    "站名": "卢氏",
                    "流量(m³/s)": 100.5  # 有数据则为数值，无数据则为None
                },
                {
                    "站名": "长水",
                    "流量(m³/s)": None  # 无数据
                },
                ...  # 共9个站点
            ]
        }
    """
    # 定义固定的10个水文站列表
    lyh_hd = ["栾川","灵口","东湾", "卢氏", "陆军","长水","龙门镇",  "宜阳", "白马寺", "黑石关"]
    # 首先获取原始数据
    auth_token = auth_token  # oauth_login()
    status_code, raw_data = get_hydrometric_station(auth_token=auth_token, station_code=station_code)
    # 初始化格式化后的数据（包含所有9个站点）
    formatted_data = {"hdsq": []}
    # 创建站点数据字典，方便查找
    station_data_map = {}
    if status_code == 200 and 'data' in raw_data and isinstance(raw_data['data'], list):
        for station in raw_data['data']:
            if str(station.get('stcd', '')).startswith('416') and station.get('stnm') in lyh_hd:
                station_data_map[station['stnm']] = station
    try:
        # 确保返回所有9个站点
        for station_name in lyh_hd:
            flow_value = None

            # 如果该站点有数据且是8点数据
            if station_name in station_data_map:
                station = station_data_map[station_name]
                timestamp = station.get('date')

                if timestamp:
                    dt = datetime.fromtimestamp(timestamp / 1000)
                    if dt.hour == 8:  # 仅当时间为8点时记录流量
                        flow_value = station.get('flow')

            formatted_data["hdsq"].append({
                "站名": station_name,
                "流量(m³/s)": flow_value
            })
        return status_code, formatted_data
    except Exception as e:
        # 出错时也返回完整结构
        for station_name in lyh_hd:
            formatted_data["hdsq"].append({
                "站名": station_name,
                "流量(m³/s)": None
            })
        return 500, {"error": f"数据处理错误: {str(e)}", "hdsq": formatted_data["hdsq"]}

def get_sk_data(auth_token):
    """
    获取河道水文站基本信息

    参数:
        station_code: 水文站代码 (String, 可选)
        auth_token: JWT认证令牌 (String)

    返回:
        (status_code, response_data) 元组
    """
    auth_token = auth_token#oauth_login()

    try:
        headers = {
            "Accept": "application/json",
            "Authorization": f"{auth_token}"
        }

        params = {}
        HYDROMETRIC_API_URL = f"{BASE_API_URL}/hydrometric/rhourrt/listLatest"
        response = requests.get(
            HYDROMETRIC_API_URL,
            params=params,
            headers=headers,
            timeout=10
        )
        response.raise_for_status()
        return response.status_code, response.json()
    except requests.exceptions.Timeout:
        return 408, {"error": "请求超时"}
    except requests.exceptions.RequestException as e:
        return 500, {"error": f"请求失败: {str(e)}"}
    except ValueError:
        return 500, {"error": "返回数据格式异常"}


def get_reservoir_properties(auth_token,ennmcd=None):
    """
    获取水库特征值信息列表

    参数:
        ennmod: 水库编码 (String, 可选)

    返回:
        (status_code, response_data) 元组
        成功时返回数据结构示例:
        {
            "code": 200,
            "data": [
                {
                    "capImmilv1": "",
                    "capNormallv1": 242.9,
                    "lvIiceMax": "",
                    "lvIFldDes": 2602.25,
                    # ...其他字段见接口文档
                }
            ]
        }
    """
    auth_token = auth_token#oauth_login()  # 获取认证token

    try:
        headers = {
            "Accept": "application/json",
            "Authorization": f"{auth_token}"
        }
        params = {}
        if ennmcd:
            params["ennmcd"] = ennmcd
        API_URL = f"{BASE_API_URL}/project/rprop/list"
        response = requests.get(
            API_URL,
            params=params,
            headers=headers,
            timeout=10
        )
        response.raise_for_status()
        return response.status_code, response.json()
    except requests.exceptions.Timeout:
        return 408, {"error": "请求超时"}
    except requests.exceptions.RequestException as e:
        return 500, {"error": f"请求失败: {str(e)}"}
    except ValueError:
        return 500, {"error": "返回数据格式异常"}


def get_reservoir_kurong(auth_token,resname):
    """
    获取水库水位-库容曲线信息
    参数:
        resname: 水库编码 (字符串，可选)
    返回:
        (status_code, response_data) 元组:
        - status_code: HTTP状态码
        - response_data: 接口返回的JSON数据或错误信息
    异常处理:
        - 处理超时、请求异常和数据格式错误
        - 返回标准化的错误信息
    示例:
        >>> status, data = get_reservoir_kurong("BDA00000121")
        >>> if status == 200:
        ...     print(data["data"][0]["capactiy"])  # 输出库容数据
    """
    # 获取认证令牌
    auth_token = auth_token
    try:
        # 构造请求头
        headers = {
            "Accept": "application/json",
            "Authorization": f"{auth_token}"  # 使用Bearer Token认证
        }

        # 构造查询参数
        params = {}
        if resname:
            params["resname"] = resname.strip()  # 去除前后空格

        # 设置API端点
        API_URL = f"{BASE_API_URL}/project/resvzv/list"

        # 发送请求（设置双超时）
        response = requests.get(
            API_URL,
            params=params,
            headers=headers,
            timeout=(3.05, 10)  # 连接超时3.05秒，读取超时10秒
        )

        # 验证响应
        response.raise_for_status()
        return response.status_code, response.json()

    except requests.exceptions.Timeout:
        return 408, {"error": "请求超时，请检查网络连接"}
    except requests.exceptions.HTTPError as e:
        return e.response.status_code, {"error": f"HTTP错误: {str(e)}"}
    except requests.exceptions.RequestException as e:
        return 500, {"error": f"请求异常: {str(e)}"}
    except ValueError as e:
        return 500, {"error": f"响应数据解析失败: {str(e)}"}

def format_reservoir_data_v1(auth_token):
    """
    获取并格式化水库数据（包含汛限水位信息）

    返回:
        dict: 格式化后的水库数据，结构如下:
        {
            "sksq": [
                {
                    "水库名称": str,
                    "水位（m）": float,
                    "蓄量（亿m³）": float,
                    "汛限水位": Optional[float],  # 允许为None
                    "相应蓄量": Optional[float],  # 允许为None
                    "超蓄水量（亿m³）": float,
                    "剩余防洪库容（亿m³）": float
                },
                ...
            ]
        }
    """
    status_code, raw_data = get_sk_data(auth_token=auth_token)
    lyh_sk = ["三门峡", "小浪底", "陆浑", "故县", "河口村"]
    formatted_data = {"sksq": []}
    try:
        if (
            status_code == 200
            and isinstance(raw_data, dict)
            and "data" in raw_data
            and isinstance(raw_data["data"], list)
        ):
            for reservoir in raw_data["data"]:
                if reservoir.get("ennm") in lyh_sk:
                    wq = reservoir.get("wq", "")
                    ennmcd = reservoir.get("ennmcd", "")
                    flood_level = None  # 默认None
                    flood_capacity = None  # 默认None
                    shengyukurong = None
                    if ennmcd:
                        prop_status, prop_data = get_reservoir_properties(auth_token=auth_token,ennmcd=ennmcd)
                        if (prop_status == 200 and prop_data.get("data")and len(prop_data["data"]) > 0):
                            # 处理汛限水位（允许None）
                            flood_level_str = prop_data["data"][0].get("lvlFldCtr", "")
                            try:
                                flood_level = float(flood_level_str) if flood_level_str else None
                            except (ValueError, TypeError):
                                flood_level = None  # 转换失败则设为None

                            # 处理相应蓄量（允许None）
                            flood_capacity_str = prop_data["data"][0].get("capCtrfldlvl", "")
                            try:
                                flood_capacity = float(flood_capacity_str) if flood_capacity_str else None
                            except (ValueError, TypeError):
                                flood_capacity = None  # 转换失败则设为None
                        kurong_status, kurong_data = get_reservoir_kurong(auth_token=auth_token,resname=ennmcd)
                        if (kurong_status == 200 and kurong_data.get("data")
                                and len(kurong_data["data"]) > 0):
                            # 使用正确的字段名（根据实际接口返回）
                            kurong = float(kurong_data['data'][-1].get("capactiy", 0))
                            shengyukurong = round(kurong - wq, 3) if kurong > wq else 0
                    # 构建格式化条目（None值不进行round计算）
                    formatted_entry = {
                        "水库名称": reservoir.get("ennm", "未知水库"),
                        "水位（m）": round(float(reservoir.get("level", 0)), 2),
                        "蓄量（亿m³）": round(float(reservoir.get("wq", 0)), 3),
                        "汛限水位（m）": round(flood_level, 2) if flood_level is not None else None,
                        "相应蓄量（亿m³）": round(flood_capacity, 3) if flood_capacity is not None else None,
                        "超蓄水量（亿m³）":round(wq-flood_capacity ,3) if flood_capacity is not None else shengyukurong,
                        "剩余防洪库容（亿m³）":shengyukurong
                    }
                    formatted_data["sksq"].append(formatted_entry)
        return formatted_data
    except Exception as e:
        print(f"数据格式化错误: {str(e)}")
        return {"sksq": []}


def format_reservoir_data(auth_token):
    """
    获取并格式化水库数据（包含汛限水位信息）
    修改：添加河名字段，简化返回字段

    返回:
        dict: 格式化后的水库数据，结构如下:
        {
            "sksq": [
                {
                    "河名": str,
                    "水库名称": str,
                    "水位（m）": float,
                    "蓄量（亿m³）": float
                },
                ...
            ]
        }
    """
    status_code, raw_data = get_sk_data(auth_token=auth_token)
    lyh_sk = ["陆浑", "故县"]
    # 水库与河名对应关系
    reservoir_to_river = {
        "陆浑": "伊河",
        "故县": "洛河"
    }

    formatted_data = {"sksq": []}
    try:
        if (
                status_code == 200
                and isinstance(raw_data, dict)
                and "data" in raw_data
                and isinstance(raw_data["data"], list)
        ):
            for reservoir in raw_data["data"]:
                reservoir_name = reservoir.get("ennm", "未知水库")
                if reservoir_name in lyh_sk:
                    # 获取对应河名
                    river_name = reservoir_to_river.get(reservoir_name, "")

                    # 构建简化后的格式化条目
                    formatted_entry = {
                        "河名": river_name,
                        "水库名称": reservoir_name,
                        "水位（m）": round(float(reservoir.get("level", 0)), 2),
                        "蓄量（亿m³）": round(float(reservoir.get("wq", 0)), 3)
                    }
                    formatted_data["sksq"].append(formatted_entry)
        return formatted_data
    except Exception as e:
        print(f"数据格式化错误: {str(e)}")
        return {"sksq": []}

import pymysql
from pymysql.cursors import DictCursor
def query_zx_reservoir_data():
    """
    连接MySQL数据库并查询特定水库的最新数据
    返回格式化后的水库数据列表
    """
    # 数据库连接配置
    db_config = {
        'host': '10.4.158.35',
        'port': 6033,
        'user': 'root',
        'password': 'awE3xDfOxX8sPduMTkoDwUBluIwJlbsi',
        'database': 'GhyDataPlatform',
        'charset': 'utf8mb4',
        'cursorclass': DictCursor
    }
    # 中型水库列表及对应的河流
    zx_sk = ['金牛岭水库', '寺河水库', '段家沟水库', '坞罗水库', '大沟口水库','洛阳市龙脖水库', '范店水库', '九龙角水库', '刘瑶水库', '陶花店水库']
    reservoir_to_river = {
        '金牛岭水库': "伊河",
        '寺河水库': "洛河",
        '段家沟水库': "洛河",
        '坞罗水库': "洛河",
        '大沟口水库': "洛河",
        '洛阳市龙脖水库': "伊河",
        '范店水库': "伊河",
        '九龙角水库': "伊河",
        '刘瑶水库': "伊河",
        '陶花店水库': "伊河"
    }
    try:
        # 建立数据库连接
        connection = pymysql.connect(**db_config)
        #print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 成功连接到数据库")
        with connection.cursor() as cursor:
            # 查询每个水库的最新数据
            formatted_results = []
            for reservoir_name in zx_sk:
                # 查询该水库的最新记录
                sql = """
                SELECT * FROM hy_res_hour_rt_all 
                WHERE ennm = %s 
                ORDER BY date DESC 
                LIMIT 1
                """
                cursor.execute(sql, (reservoir_name,))
                latest_data = cursor.fetchone()
                #print("latest_data", latest_data)
                if latest_data:
                    # 格式化数据
                    formatted_entry = {
                        "河名": reservoir_to_river.get(reservoir_name, "未知"),
                        "水库名称": reservoir_name,
                        "水位（m）": round(float(latest_data.get("level", 0)), 2),
                        "蓄量（亿m³）": round(float(latest_data.get("wq", 0)), 3)  # 假设原单位是万m³，转换为亿m³
                    }
                    formatted_results.append(formatted_entry)
            #print(f"共查询到 {len(formatted_results)} 座水库的最新数据")
            return formatted_results
    except pymysql.MySQLError as e:
        print(f"数据库连接或查询失败: {e}")
        return []
    finally:
        if 'connection' in locals() and connection.open:
            connection.close()
            #print("数据库连接已关闭")

def get_hydrometric_dayrt_list(auth_token,hysta=None, start_date=None, end_date=None):
    """
    获取水文站日均水情信息列表

    参数:
        hysta: 水文站代码 (String, 可选)
        start_date: 开始日期 (String, 格式: YYYY-MM-DD, 默认昨天)
        end_date: 结束日期 (String, 格式: YYYY-MM-DD, 默认今天)

    返回:
        (status_code, response_data) 元组
    """
    auth_token = auth_token#oauth_login()  # 获取认证token
    try:
        # 设置默认日期
        today = datetime.now()
        yesterday = today - timedelta(days=1)
        if not start_date:
            start_date = yesterday.strftime('%Y-%m-%d')
        if not end_date:
            end_date = today.strftime('%Y-%m-%d')
        headers = {
            "Accept": "application/json",
            "Authorization": f"{auth_token}"
        }
        params = {
            "startDate": start_date,
            "endDate": end_date
        }
        if hysta:
            params["hysta"] = hysta
        API_URL = f"{BASE_API_URL}/hydrometric/dayrt/list"
        response = requests.get(
            API_URL,
            params=params,
            headers=headers,
            timeout=10
        )
        response.raise_for_status()
        return response.status_code, response.json()
    except requests.exceptions.Timeout:
        return 408, {"error": "请求超时"}
    except requests.exceptions.RequestException as e:
        return 500, {"error": f"请求失败: {str(e)}"}
    except ValueError:
        return 500, {"error": "返回数据格式异常"}

def get_access_token(
        username: str = "admin1",
        password: str = "Yrec!@#2025",
        client_id: str = "e5cd7e4891bf95d1d19206ce24a7b32e",
        grant_type: str = "password",
        base_url: str = "http://10.4.158.35:8091"  # 根据实际环境修改
):
    """
    获取认证Token

    参数:
        username: 用户名 (默认示例值)
        password: 密码 (默认示例值)
        client_id: 客户端ID (默认示例值)
        grant_type: 授权类型 (固定password)
        base_url: 接口基础地址

    返回:
        str: 成功返回access_token，失败返回None
    """
    url = f"{base_url}/auth/login"
    payload = {
        "username": username,
        "password": password,
        "clientId": client_id,
        "grantType": grant_type
    }
    try:
        response = requests.post(
            url=url,
            json=payload,  # 使用json参数自动设置Content-Type为application/json
            timeout=10
        )
        response.raise_for_status()

        data = response.json()
        if data.get("code") == 200:
            return data["data"]["access_token"]
        else:
            print(f"登录失败: {data.get('msg')}")
            return None

    except requests.exceptions.RequestException as e:
        print(f"请求异常: {str(e)}")
        return None
    except (KeyError, ValueError) as e:
        print(f"数据解析错误: {str(e)}")
        return None

# 使用示例
def get_weather_warning(auth_token: str, timeout: int = 10):
    """
    获取天气预警信息并生成描述文本（带去重功能）

    参数:
        auth_token (str): 认证Token
        timeout (int): 请求超时时间(秒)，默认10秒

    返回:
        dict: 包含响应状态和数据的字典，格式:
        {
            "status_code": int,    # HTTP状态码
            "data": dict | str,    # 成功时为JSON数据，失败时为错误信息
            "description": str     # 生成的预警描述文本
        }
    """
    url = "http://10.4.158.35:8091/pre/getWeatherWarning"
    try:
        response = requests.get(
            url=url,
            headers={
                "ClientId": "e5cd7e4891bf95d1d19206ce24a7b32e",
                "Authorization": f"Bearer {auth_token}"
            },
            timeout=timeout
        )
        response.raise_for_status()
        result = response.json()

        # 初始化预警分类字典（使用集合自动去重）
        warning_levels = {
            "红色": set(),
            "橙色": set(),
            "黄色": set(),
            "蓝色": set()
        }

        # 处理数据并生成描述
        description = "根据伊洛河流域县区气象台发布的预警信息"

        if result.get("code") == 200 and result.get("data"):
            # 提取并分类预警信息（自动去重）
            for warning in result["data"]:
                city_name = warning.get("cityName", "")
                warning_name = warning.get("warningName", "")

                # 提取县区名称（改进的提取逻辑）
                county = extract_county_name(city_name)

                # 分类预警级别
                if county:  # 只有有效县区名称才添加
                    if "红色" in warning_name:
                        warning_levels["红色"].add(county)
                    elif "橙色" in warning_name:
                        warning_levels["橙色"].add(county)
                    elif "黄色" in warning_name:
                        warning_levels["黄色"].add(county)
                    elif "蓝色" in warning_name:
                        warning_levels["蓝色"].add(county)

            # 构建描述文本
            parts = []
            for level, counties in warning_levels.items():
                if counties:
                    # 将集合转为列表并排序，确保输出顺序一致
                    sorted_counties = sorted(list(counties))
                    counties_str = "、".join(sorted_counties)
                    # 单数/复数处理
                    if len(sorted_counties) > 1:
                        parts.append(f"{counties_str}等{len(sorted_counties)}个县区发布了暴雨{level}预警")
                    else:
                        parts.append(f"{counties_str}发布了暴雨{level}预警")

            if parts:
                description += "，" + "，".join(parts) + "。"
            else:
                description += "当前无暴雨预警。"
        else:
            description += "当前无暴雨预警。"

        return {
            "status_code": response.status_code,
            "data": result,
            "description": description
        }

    except requests.exceptions.Timeout:
        return {
            "status_code": 408,
            "data": "请求超时",
            "description": "获取预警信息超时"
        }
    except requests.exceptions.RequestException as e:
        return {
            "status_code": 500,
            "data": f"请求失败: {str(e)}",
            "description": "获取预警信息失败"
        }
    except ValueError:
        return {
            "status_code": 500,
            "data": "返回数据格式异常",
            "description": "预警信息解析错误"
        }

def get_rain_analysis(auth_token: str, timeout: int = 10):
    """
    获取降雨分析信息并生成描述文本（自动根据当前时间设置时间段）

    参数:
        auth_token (str): 认证Token
        timeout (int): 请求超时时间(秒)，默认10秒

    返回:
        dict: 包含响应状态和数据的字典，格式:
        {
            "status_code": int,    # HTTP状态码
            "data": dict | str,    # 成功时为JSON数据，失败时为错误信息
            "description": str     # 生成的降雨分析描述文本
        }
    """
    url = "http://10.4.158.35:8091/rainfall/chy/rainAnalysis"

    # 获取当前时间
    now = datetime.now()
    current_hour = now.hour

    # 自动判断开始时间和结束时间
    if current_hour < 8:
        # 当前时间小于 8 点，取昨天 08:00 到当前时间
        start_time = (now - timedelta(days=1)).replace(hour=8, minute=0, second=0, microsecond=0)
    else:
        # 当前时间大于等于 8 点，取今天 08:00 到当前时间
        start_time = now.replace(hour=8, minute=0, second=0, microsecond=0)

    end_time = now

    # 格式化时间字符串（ISO格式）
    start_time_str = start_time.strftime("%Y-%m-%d %H:%M:%S")
    end_time_str = end_time.strftime("%Y-%m-%d %H:%M:%S")
    # start_time_str = "2025-05-04 08:00:00"
    # end_time_str = "2025-05-05 01:00:00"
    full_url = f"{url}?startTime={start_time_str}&endTime={end_time_str}"
    print("full_url：",full_url)
    try:
        response = requests.get(
            url=full_url,
            headers={
                "ClientId": "e5cd7e4891bf95d1d19206ce24a7b32e",
                "Authorization": f"Bearer {auth_token}"
            },
            timeout=timeout
        )
        response.raise_for_status()

        result = response.json()
        print("result:",result)
        print("type(result[data])",type(result['data']))
        # 初始化描述文本
        return {
            "status_code": response.status_code,
            "data":result['data']
        }

    except requests.exceptions.Timeout:
        return {
            "status_code": 408,
            "data": "访问接口超时"
        }
    except requests.exceptions.RequestException as e:
        return {
            "status_code": 500,
            "data": f"请求失败: {str(e)}"
        }
    except ValueError:
        return {
            "status_code": 500,
            "data": "返回数据格式异常"
        }

import os
from shapely.geometry import Point
from matplotlib.colors import Normalize
from matplotlib.cm import ScalarMappable
from matplotlib.colors import ListedColormap, BoundaryNorm
from matplotlib.patches import Patch
import geopandas as gpd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D
def plot_yuliangmian(
    rain_geojson_result,
    max_rainfall_station,
    basin_geojson_path="data/geojson/洛河流域.json",
    water_geojson_path="data/geojson/WTRIVRL25_洛河流域.json"
):
    """
    绘制洛河流域边界、雨量分布面数据、水系（河流）、最大降雨站点

    参数:
        rain_geojson_result (dict): get_rain_analysis 返回结果，可能为 None
        max_rainfall_station (dict): get_max_rainfall_station 返回的最大降雨站点信息
        basin_geojson_path (str): 流域 GeoJSON 文件路径
        water_geojson_path (str): 水系 GeoJSON 文件路径
    """

    # 设置中文字体支持
    #plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'WenQuanYi Zen Hei', 'Noto Sans CJK SC']
    plt.rcParams['font.sans-serif'] = ['WenQuanYi Micro Hei',  # 文泉驿微米黑
        'Noto Sans CJK JP',  # Noto 日文（兼容简体中文）
        'Noto Serif CJK JP'  # Noto 衬线体
    ]
    plt.rcParams['axes.unicode_minus'] = False
    plt.rcParams['axes.labelsize'] = 1  # 隐藏默认标签

    # 提取雨量面数据（兼容 None）
    data = None
    if rain_geojson_result and isinstance(rain_geojson_result.get("data"), dict):
        data = rain_geojson_result["data"]
    else:
        print("未获取到有效的雨量面数据，将不绘制雨量分布。")

    # 加载洛河流域边界 GeoJSON（如果存在）
    basin_gdf = None
    if os.path.exists(basin_geojson_path):
        try:
            basin_gdf = gpd.read_file(basin_geojson_path)
        except Exception as e:
            print(f"读取洛河流域 GeoJSON 失败: {e}")
    else:
        print(f"未找到洛河流域 GeoJSON 文件: {basin_geojson_path}")

    # 加载水系数据（河流线状数据）
    water_gdf = None
    if os.path.exists(water_geojson_path):
        try:
            water_gdf = gpd.read_file(water_geojson_path).to_crs("EPSG:4326")
        except Exception as e:
            print(f"读取水系 GeoJSON 失败: {e}")
    else:
        print(f"未找到水系 GeoJSON 文件: {water_geojson_path}")

    # 创建绘图
    fig, ax = plt.subplots(figsize=(12, 8))
    ax.set_facecolor('white')

    # 绘制洛河流域边界（蓝色）
    if basin_gdf is not None:
        basin_gdf.boundary.plot(ax=ax, color='lightgray', linewidth=1)

    # 定义降雨量等级与颜色
    bins = [0, 0.1, 10, 25, 50, 100, 250, float('inf')]
    colors = ['white', '#eaffcc', '#b3f7a1', '#88d8ff', '#007fff', '#ff3300', '#9900cc']
    labels = ['0 - 0.1 mm', '0.1 - 10 mm', '10 - 25 mm', '25 - 50 mm',
              '50 - 100 mm', '100 - 250 mm', '>250 mm']

    # 添加图例元素（Patch + 红点）
    legend_elements = [
        Patch(facecolor=color, edgecolor='gray', label=label)
        for color, label in zip(colors, labels)
    ]

    # 如果没有有效雨量数据，在图例顶部加一个提示项
    # if data is None or 'features' not in data:
    #     legend_elements.insert(0, Patch(facecolor='lightgray', edgecolor='black', label='无雨量数据'))

    # 如果有最大降雨站点，加入图例
    # if max_rainfall_station:
    #     legend_elements.append(
    #         Line2D([0], [0], marker='o', color='none', label='最大降雨站点',
    #                markerfacecolor='red', markersize=10, linestyle='')
    #     )
    # 解析并绘制雨量分布面数据（如有）
    if data and 'features' in data:
        try:
            rain_gdf = gpd.GeoDataFrame.from_features(data['features'])
            if 'value' in rain_gdf.columns:
                cmap = ListedColormap(colors)
                norm = BoundaryNorm(boundaries=bins, ncolors=len(colors))
                rain_gdf.plot(
                    ax=ax,
                    column='value',
                    cmap=cmap,
                    norm=norm,
                    edgecolor='gray',
                    linewidth=0.5,
                    legend=False,
                    zorder=2
                )
        except Exception as e:
            print(f"解析雨量面数据失败: {e}")

    # 绘制水系（河流）
    if water_gdf is not None:
        water_gdf.plot(ax=ax, color="lightblue", linewidth=0.5, zorder=1, label='河流网络')
        # river_gdf.plot(ax=ax, color="#ADD8E6", linewidth=0.8, label="河流网络")
        # legend_elements.append(
        #     Line2D([0], [0], color='darkblue', lw=1.5, label='河流网络')
        # )

    # 绘制最大降雨站点（红点 + 站名标注）
    if max_rainfall_station:
        point = Point(max_rainfall_station['lgtd'], max_rainfall_station['lttd'])
        point_gdf = gpd.GeoDataFrame([{'geometry': point, 'name': max_rainfall_station['stnm']}],
                                     crs="EPSG:4326")
        point_gdf.plot(ax=ax, color='red', marker='o', markersize=70, zorder=3)

        # 添加站名和雨量标注
        ax.text(max_rainfall_station['lgtd'], max_rainfall_station['lttd'],
                f" {max_rainfall_station['stnm']}\n{max_rainfall_station['rf']}mm",
                fontsize=11, ha='left', va='bottom', color='darkred')

    # 设置标题
    ax.set_title("伊洛河流域面雨量", fontsize=14)
    ax.set_axis_off()  # 不显示坐标轴

    # 添加图例（放在右下角）
    ax.legend(handles=legend_elements, loc='lower right', fontsize=9, title="降雨量等级")

    plt.tight_layout()

    # 自动创建输出文件夹并保存图片
    today = datetime.now().strftime("%Y-%m-%d")
    output_folder = os.path.join("data", "yuan_data", "4", "yubao", today)
    os.makedirs(output_folder, exist_ok=True)
    output_path = os.path.join(output_folder, "yuliang.png")
    if os.path.exists(output_path):
        os.remove(output_path)
        print(f"🗑️ 已删除旧文件: {output_path}")
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    print(f"✅ 图片已保存至: {output_path}")
    # 关闭图像以释放内存
    plt.close()


def extract_county_name(city_name: str) -> str:
    """
    从cityName字段中提取县区名称

    参数:
        city_name: 原始城市名称字符串（如"河南省洛阳市嵩县气象台发布..."）

    返回:
        提取后的县区名称（如"嵩县"），提取失败返回空字符串
    """
    if not city_name:
        return ""

    # 尝试按"气象台"分割
    parts = city_name.split("气象台")
    if len(parts) > 0:
        # 尝试按省市区分割
        address_parts = parts[0].split("省")[-1].split("市")
        if len(address_parts) > 1:
            return address_parts[-1]
        return parts[0]
    return ""

def get_formatted_jlyb_data(auth_token, dateTime=None):
    """
    获取格式化后的降雨预报数据

    参数:
    dateTime -- 查询时间字符串，格式为'YYYY-MM-DD HH:MM:SS'，默认为当天8点

    返回:
    格式化后的降雨数据列表，或None(如果出错)
    """
    # API基础URL
    base_url = "http://10.4.158.35:8091/pre/getPreRainDataByTime"

    # 设置默认时间为当天8点
    if dateTime is None:
        today = datetime.now().strftime("%Y-%m-%d")
        dateTime = f"{today} 08:00:00"

    try:
        # 发送GET请求带参数
        params = {"dateTime": dateTime}
        response = requests.get(base_url,  headers={
                "ClientId": "e5cd7e4891bf95d1d19206ce24a7b32e",
                "Authorization": f"Bearer {auth_token}"
            }, params=params)
        response.raise_for_status()  # 检查请求是否成功

        # 解析JSON响应
        data = response.json()

        # 检查返回状态码
        if data.get("code") != 200:
            raise ValueError(f"API返回错误: {data.get('msg', '未知错误')}")

        # 转换数据格式
        formatted_data = []
        for item in data.get("data", []):
            formatted_item = {
                "区域": item.get("stationName", ""),
                "0-24小时": item.get("value24", ""),
                "24-48小时": item.get("value48", ""),
                "48-72小时": item.get("value72", "")
            }
            formatted_data.append(formatted_item)

        return formatted_data

    except requests.exceptions.RequestException as e:
        print(f"请求API时出错: {e}")
        return None
    except ValueError as e:
        print(f"数据处理错误: {e}")
        return None
    except Exception as e:
        print(f"发生未知错误: {e}")
        return None


def get_ddfad_data(auth_token, base_url="http://10.4.158.35:8091"):
    """
        调度方案单数据
    """
    # API基础URL

    try:
        headers = {
            "Authorization": f"Bearer {auth_token}",
            "ClientId": "e5cd7e4891bf95d1d19206ce24a7b32e"
        }

        url = f"{base_url}/preSch/getSch"
        # print("ddfad:", url, headers)
        response = requests.get(
            url,
            headers=headers,
            timeout=10
        )

        response.raise_for_status()  # 检查请求是否成功

        schId = response.json()["data"][0]["id"]
        isRecommend = response.json()["data"][0]["isRecommend"]
        logger.info(f"get_ddfad_data schId: {schId}, isRecommend: {isRecommend}")
        xurl = f"{base_url}/preSch/getSchDataBySchId?schId={schId}"
        print("ddfad:", xurl)
        # print("ddfad:", url, headers)
        response = requests.get(
            xurl,
            headers=headers,
            timeout=10
        )
        response.raise_for_status()  # 检查请求是否成功
        return response.status_code, response.json()

    except requests.exceptions.Timeout:
        return 408, {"error": "请求超时"}
    except requests.exceptions.RequestException as e:
        return 500, {"error": f"请求失败: {str(e)}"}
    except ValueError:
        return 500, {"error": "返回数据格式异常"}

def get_tufang_data(auth_token, base_url="http://10.4.158.35:8091"):
    """
        获取防汛土方数据
    """
    # API基础URL

    try:
        headers = {
            "Authorization": f"Bearer {auth_token}",
            "ClientId": "e5cd7e4891bf95d1d19206ce24a7b32e"
        }

        url = f"{base_url}/preSch/getSch"
        # print("ddfad:", url, headers)
        response = requests.get(
            url,
            headers=headers,
            timeout=10
        )

        response.raise_for_status()  # 检查请求是否成功

        schId = response.json()["data"][0]["id"]
        isRecommend = response.json()["data"][0]["isRecommend"]
        logger.info(f"get_tufang_data schId: {schId}, isRecommend: {isRecommend}")
        xurl = f"{base_url}/dispatch/getDfOverLevel?schId={schId}"
        print("get_tufang_data:", xurl)
        # print("ddfad:", url, headers)
        response = requests.get(
            xurl,
            headers=headers,
            timeout=10
        )
        response.raise_for_status()  # 检查请求是否成功
        return response.status_code, response.json()

    except requests.exceptions.Timeout:
        return 408, {"error": "请求超时"}
    except requests.exceptions.RequestException as e:
        return 500, {"error": f"请求失败: {str(e)}"}
    except ValueError:
        return 500, {"error": "返回数据格式异常"}
# if __name__ == '__main__':
#     import time
#     import matplotlib.pyplot as plt
#     res = search_fragpacks("雨水情信息")
#     print(res)
#     # 记录开始时间
#     r = yautils.excel_to_dict("../../mainapp/media/ddfa/3/2025-03-03.xlsx")
#     skMapData, swMapData, date_list = r
#     # pprint.pprint(r)
#     start_time = time.time()
#     # 调用函数并获取结果
#     res = yautils.skddjy_new("../../mainapp/media/ddfa/3/2025-03-03.xlsx")
#     ddjy_list = []
#     # 遍历返回的结果字典
#     for key, value in res.items():
#         # 处理每个水库的出流量
#         ckll = process_outflow(value, date_list)
#         ddjy_list.append(f"{key}水库：{ckll}")
#     # 将所有水库的调度建议合并为一个字符串
#     ddjy = "\n".join(ddjy_list)
#     # 记录结束时间
#     end_time = time.time()
#     # 计算并打印执行时间
#     execution_time = end_time - start_time
#     print(f"水库调度建议：\n{ddjy}")
#     print(f"代码执行时间：{execution_time:.4f} 秒")
#     plt.rcParams['font.sans-serif'] = ['SimHei']  # 使用黑体
#     plt.rcParams['axes.unicode_minus'] = False  # 解决负号显示问题
#
#     # 示例数据
#     data = {
#         "水库名称": ["故县水库", "陆浑水库", "河口村水库", "三门峡水库", "小浪底水库"],
#         "开始时间": ["2021-10-27 08:00:00"] * 5,
#         "结束时间": ["2021-10-27 22:00:00"] * 5,
#         "流量": [1000, 900, 4000, 500, 4500],  # 流量值，三门峡为敞泄（假设为0）
#     }
#
#     # 转换为 DataFrame
#     df = pd.DataFrame(data)
#
#     # 将时间转换为 datetime 格式
#     df["开始时间"] = pd.to_datetime(df["开始时间"])
#     df["结束时间"] = pd.to_datetime(df["结束时间"])
#
#     # 创建图表
#     plt.figure(figsize=(12, 6))
#
#     # 绘制每个水库的调度时间段
#     for i, row in df.iterrows():
#         # 绘制时间段
#         plt.plot([row["开始时间"], row["结束时间"]], [row["流量"], row["流量"]],
#                  marker="o", label=row["水库名称"], linewidth=2)
#         # # 标注流量值
#         # plt.text(row["开始时间"], row["流量"], f"{row['流量']} m³/s",
#         #          va="bottom", ha="right", fontsize=10, color="blue")
#
#     # 设置图表属性
#     plt.xlabel("时间", fontsize=12)
#     plt.ylabel("出库流量 (立方米/秒)", fontsize=12)
#     plt.title("水库调度方式时间序列图", fontsize=14)
#     plt.grid(True, linestyle="--", alpha=0.6)  # 网格线
#     plt.legend(loc="upper right", fontsize=10)  # 图例
#
#     # 设置 X 轴时间格式
#     plt.gca().xaxis.set_major_formatter(plt.matplotlib.dates.DateFormatter('%Y-%m-%d %H:%M:%S'))
#     plt.gcf().autofmt_xdate()  # 自动旋转 X 轴标签
#
#     # 显示图表
#     plt.tight_layout()
#     plt.show()